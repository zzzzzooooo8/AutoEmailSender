from __future__ import annotations

import csv
import io
from html import unescape
import re
from dataclasses import dataclass
from typing import Any
from zipfile import BadZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.exceptions import InvalidFileException

from app.schemas.professor import ProfessorUpsertPayload
from app.services.professor_field_normalization import normalize_recent_papers


PROFESSOR_TEMPLATE_COLUMNS = [
    "name",
    "email",
    "title",
    "university",
    "school",
    "department",
    "research_direction",
    "recent_papers",
    "profile_url",
    "source_url",
]

PROFESSOR_TEMPLATE_HELP_LINES = [
    "# 导师导入模板",
    "# 从字段名下一行开始填写；说明行和示例行可以保留，系统导入时会自动忽略",
    "# 必填字段：name, email",
    "# name：导师姓名，必填。示例：张明远",
    "# email：导师邮箱，必填，必须是邮箱格式。示例：zhang@example.edu",
    "# title：导师职称。示例：教授",
    "# university：学校名称。示例：示例大学",
    "# school：学院名称。示例：人工智能学院",
    "# department：院系或系所。示例：计算机科学系",
    "# research_direction：研究方向，多个方向用中文分号 ； 分隔。示例：大语言模型；智能体；信息抽取",
    "# recent_papers：近期论文，多篇用 | 分隔；最多保留前 8 篇。示例：Paper A|Paper B",
    "# profile_url：导师主页链接。示例：https://example.edu/zhang",
    "# source_url：数据来源链接。示例：https://example.edu/faculty",
]

PROFESSOR_TEMPLATE_EXAMPLE_ROW = [
    "示例：张明远",
    "zhang@example.edu",
    "教授",
    "示例大学",
    "人工智能学院",
    "计算机科学系",
    "大语言模型；智能体；信息抽取",
    "Paper A|Paper B",
    "https://example.edu/zhang",
    "https://example.edu/faculty",
]

SUPPORTED_IMPORT_EXTENSIONS = {".csv", ".xlsx"}
EMAIL_LOCAL_PATTERN = re.compile(r"^[A-Za-z0-9._%+-]+$")
EMAIL_DOMAIN_LABEL_PATTERN = re.compile(r"^[A-Za-z0-9](?:[A-Za-z0-9-]{0,61}[A-Za-z0-9])?$")
EMAIL_FULLWIDTH_TRANSLATION = str.maketrans(
    {
        "＠": "@",
        "．": ".",
        "。": ".",
        "﹒": ".",
        "｡": ".",
        "（": "(",
        "）": ")",
        "［": "[",
        "］": "]",
        "【": "[",
        "】": "]",
        "｛": "{",
        "｝": "}",
    }
)
EMAIL_INVISIBLE_PATTERN = re.compile(r"[\u200b\u200c\u200d\ufeff]")
EMAIL_CHINESE_EMAIL_SYMBOL_PATTERN = re.compile(r"邮箱符号")
EMAIL_AT_PATTERN = re.compile(r"[\(\[\s]*at[\)\]\s]*", re.IGNORECASE)
EMAIL_DOT_PATTERN = re.compile(r"[\(\[\s]*dot[\)\]\s]*", re.IGNORECASE)
EMAIL_CHINESE_DOT_PATTERN = re.compile(r"(?<=[A-Za-z0-9])\s*点\s*(?=[A-Za-z0-9])")
TITLE_SPLIT_PATTERN = re.compile(r"[、，,/／|｜；;\s]+")
ALLOWED_TITLES = (
    "教授",
    "副教授",
    "助理教授",
    "讲师",
    "研究员",
    "副研究员",
    "助理研究员",
    "特聘研究员",
)
TITLE_PRIORITY = {title: index for index, title in enumerate(ALLOWED_TITLES)}


@dataclass(slots=True)
class ParsedProfessorImport:
    data: dict[str, Any]
    failed_count: int


def normalize_professor_email(email: str | None) -> str | None:
    if email is None:
        return None

    cleaned = _normalize_email_text(str(email))
    if not cleaned:
        return None
    if cleaned.count("@") != 1:
        return cleaned

    local_part, domain = cleaned.split("@", 1)
    normalized_domain = re.sub(r"\.{2,}", ".", domain)
    normalized_domain = normalized_domain.strip(".")
    return f"{local_part}@{normalized_domain}" if normalized_domain else cleaned


def is_valid_professor_email(email: str) -> bool:
    cleaned = email.strip()
    if cleaned.count("@") != 1:
        return False

    local_part, domain = cleaned.split("@", 1)
    if not local_part or not domain:
        return False
    if not EMAIL_LOCAL_PATTERN.fullmatch(local_part):
        return False

    labels = domain.split(".")
    if len(labels) < 2:
        return False
    if not all(EMAIL_DOMAIN_LABEL_PATTERN.fullmatch(label) for label in labels):
        return False
    return labels[-1].isalpha() and len(labels[-1]) >= 2


def _normalize_email_text(value: str) -> str:
    normalized = unescape(value).strip().lower()
    normalized = normalized.translate(EMAIL_FULLWIDTH_TRANSLATION)
    normalized = EMAIL_INVISIBLE_PATTERN.sub("", normalized)
    normalized = EMAIL_CHINESE_EMAIL_SYMBOL_PATTERN.sub("@", normalized)
    normalized = EMAIL_AT_PATTERN.sub("@", normalized)
    normalized = EMAIL_DOT_PATTERN.sub(".", normalized)
    normalized = EMAIL_CHINESE_DOT_PATTERN.sub(".", normalized)
    normalized = re.sub(r"\s*@\s*", "@", normalized)
    normalized = re.sub(r"\s*\.\s*", ".", normalized)
    normalized = re.sub(r"\s+", "", normalized)
    return normalized



def normalize_professor_title(title: str | None) -> str | None:
    if title is None:
        return None

    normalized = str(title).strip()
    if not normalized:
        return None

    matched_titles = []
    seen = set()
    for segment in TITLE_SPLIT_PATTERN.split(normalized):
        candidate = segment.strip()
        if not candidate or candidate in seen:
            continue
        if candidate in TITLE_PRIORITY:
            matched_titles.append(candidate)
            seen.add(candidate)

    if not matched_titles:
        return None

    return min(matched_titles, key=lambda item: TITLE_PRIORITY[item])


def normalize_professor_payload(payload: ProfessorUpsertPayload) -> dict[str, Any]:
    return {
        "name": payload.name.strip(),
        "email": normalize_professor_email(payload.email) or "",
        "title": normalize_professor_title(payload.title),
        "university": payload.university,
        "school": payload.school,
        "department": payload.department,
        "research_direction": payload.research_direction,
        "recent_papers": payload.recent_papers,
        "profile_url": payload.profile_url,
        "source_url": payload.source_url,
    }


def build_professor_template(format_name: str) -> tuple[bytes, str, str]:
    normalized = format_name.lower()
    if normalized == "csv":
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        for line in PROFESSOR_TEMPLATE_HELP_LINES:
            writer.writerow([line])
        writer.writerow(PROFESSOR_TEMPLATE_COLUMNS)
        writer.writerow(PROFESSOR_TEMPLATE_EXAMPLE_ROW)
        content = buffer.getvalue().encode("utf-8-sig")
        return content, "text/csv; charset=utf-8", "professors_import_template.csv"

    if normalized == "xlsx":
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Professors"
        help_fill = PatternFill("solid", fgColor="F5F5F4")
        header_fill = PatternFill("solid", fgColor="E7E5E4")

        for line in PROFESSOR_TEMPLATE_HELP_LINES:
            sheet.append([line])
            row_index = sheet.max_row
            sheet.merge_cells(
                start_row=row_index,
                start_column=1,
                end_row=row_index,
                end_column=len(PROFESSOR_TEMPLATE_COLUMNS),
            )
            cell = sheet.cell(row=row_index, column=1)
            cell.alignment = Alignment(wrap_text=True)
            cell.fill = help_fill

        sheet.append(PROFESSOR_TEMPLATE_COLUMNS)
        header_row = sheet.max_row
        for index, column in enumerate(PROFESSOR_TEMPLATE_COLUMNS, start=1):
            cell = sheet.cell(row=header_row, column=index)
            cell.value = column
            cell.font = Font(bold=True)
            cell.fill = header_fill
            sheet.column_dimensions[chr(64 + index)].width = 22
        sheet.append(PROFESSOR_TEMPLATE_EXAMPLE_ROW)
        sheet.freeze_panes = f"A{header_row + 1}"

        buffer = io.BytesIO()
        workbook.save(buffer)
        return (
            buffer.getvalue(),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "professors_import_template.xlsx",
        )

    raise ValueError("仅支持 csv 或 xlsx 模板")


def parse_professor_import_file(filename: str, content: bytes) -> ParsedProfessorImport:
    suffix = _resolve_suffix(filename)
    if suffix == ".csv":
        rows = _parse_csv_rows(content)
    else:
        rows = _parse_xlsx_rows(content)

    deduplicated: dict[str, dict[str, Any]] = {}
    failed_count = 0
    for row in rows:
        if _should_skip_import_row(row):
            continue
        normalized = _normalize_import_row(row)
        if normalized is None:
            failed_count += 1
            continue
        deduplicated[normalized["email"]] = normalized

    return ParsedProfessorImport(
        data=deduplicated,
        failed_count=failed_count,
    )


def _resolve_suffix(filename: str) -> str:
    lower_name = filename.lower()
    for suffix in SUPPORTED_IMPORT_EXTENSIONS:
        if lower_name.endswith(suffix):
            return suffix
    raise ValueError("仅支持导入 csv 或 xlsx 文件")


def _parse_csv_rows(content: bytes) -> list[dict[str, Any]]:
    try:
        decoded = content.decode("utf-8-sig")
    except UnicodeDecodeError as error:
        raise ValueError("CSV 文件请使用 UTF-8 编码") from error

    reader = csv.reader(io.StringIO(decoded))
    return _parse_tabular_rows(list(reader))


def _parse_xlsx_rows(content: bytes) -> list[dict[str, Any]]:
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except (BadZipFile, InvalidFileException, OSError) as error:
        raise ValueError("XLSX 文件无法读取") from error
    sheet = workbook.active
    return _parse_tabular_rows(list(sheet.iter_rows(values_only=True)))


def _parse_tabular_rows(rows: list[list[Any] | tuple[Any, ...]]) -> list[dict[str, Any]]:
    if not rows:
        raise ValueError("导入文件为空")

    header_index = _find_header_row_index(rows)
    header_names = [str(cell).strip() if cell is not None else "" for cell in rows[header_index]]
    _validate_columns(header_names)

    parsed_rows: list[dict[str, Any]] = []
    for row in rows[header_index + 1 :]:
        if _is_help_row(row):
            continue
        parsed_rows.append(
            {
                header_names[index]: row[index] if index < len(row) else None
                for index in range(len(header_names))
            }
        )
    return parsed_rows


def _find_header_row_index(rows: list[list[Any] | tuple[Any, ...]]) -> int:
    for index, row in enumerate(rows):
        normalized = [str(cell).strip() if cell is not None else "" for cell in row]
        if all(column in normalized for column in PROFESSOR_TEMPLATE_COLUMNS):
            return index
    raise ValueError(f"导入文件缺少必要列：{', '.join(PROFESSOR_TEMPLATE_COLUMNS)}")


def _validate_columns(columns: list[str] | tuple[str, ...] | None) -> None:
    if columns is None:
        raise ValueError("导入文件缺少表头")

    normalized = [str(item).strip() for item in columns]
    missing = [column for column in PROFESSOR_TEMPLATE_COLUMNS if column not in normalized]
    if missing:
        raise ValueError(f"导入文件缺少必要列：{', '.join(missing)}")


def _normalize_import_row(row: dict[str, Any]) -> dict[str, Any] | None:
    raw_values = {key: _clean_cell_value(row.get(key)) for key in PROFESSOR_TEMPLATE_COLUMNS}
    if not any(raw_values.values()):
        return None

    name = raw_values["name"]
    email = normalize_professor_email(raw_values["email"]) or ""
    if not name or not email or not is_valid_professor_email(email):
        return None

    return {
        "name": name,
        "email": email,
        "title": normalize_professor_title(raw_values["title"]),
        "university": raw_values["university"],
        "school": raw_values["school"],
        "department": raw_values["department"],
        "research_direction": raw_values["research_direction"],
        "recent_papers": _parse_recent_papers(raw_values["recent_papers"]),
        "profile_url": raw_values["profile_url"],
        "source_url": raw_values["source_url"],
    }


def _should_skip_import_row(row: dict[str, Any]) -> bool:
    raw_values = {key: _clean_cell_value(row.get(key)) for key in PROFESSOR_TEMPLATE_COLUMNS}
    if not any(raw_values.values()):
        return True
    name = raw_values["name"] or ""
    return name.startswith("#") or name.startswith("示例：")


def _is_help_row(row: list[Any] | tuple[Any, ...]) -> bool:
    first_cell = _clean_cell_value(row[0]) if row else None
    return bool(first_cell and first_cell.startswith("#"))


def _clean_cell_value(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _parse_recent_papers(value: str | None) -> list[str]:
    return normalize_recent_papers(value)
