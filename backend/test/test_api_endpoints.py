from __future__ import annotations

import asyncio
import io
import json
import os
import sqlite3
import subprocess
import sys
import tempfile
import unittest
from datetime import UTC, datetime, timedelta
from pathlib import Path
from unittest.mock import AsyncMock, patch

from openpyxl import Workbook, load_workbook
from fastapi.testclient import TestClient


BACKEND_DIR = Path(__file__).resolve().parents[1]
HEAD_REVISION = "b9d1e3f4a6c7"


class ApiEndpointTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp_dir = tempfile.TemporaryDirectory()
        self.db_path = Path(self.temp_dir.name) / "api_test.db"
        os.environ["DATABASE_URL"] = f"sqlite+aiosqlite:///{self.db_path.as_posix()}"
        os.environ["ENABLE_BACKGROUND_WORKERS"] = "0"
        self._run_alembic_upgrade()

        from app.core.config import get_settings
        from app.core.database import dispose_engine, get_engine, get_session_factory
        from main import create_app

        get_settings.cache_clear()
        if get_engine.cache_info().currsize:
            asyncio.run(dispose_engine())
        get_session_factory.cache_clear()
        get_settings.cache_clear()

        self.client = TestClient(create_app())

    def tearDown(self) -> None:
        self.client.close()
        from app.core.config import get_settings
        from app.core.database import dispose_engine, get_engine, get_session_factory

        if get_engine.cache_info().currsize:
            asyncio.run(dispose_engine())
        get_session_factory.cache_clear()
        get_settings.cache_clear()
        os.environ.pop("DATABASE_URL", None)
        os.environ.pop("ENABLE_BACKGROUND_WORKERS", None)
        self.temp_dir.cleanup()

    def test_identity_and_llm_connectivity_endpoints(self) -> None:
        identity_id = self._create_identity(with_imap=False)
        llm_id = self._create_llm()
        identities = self.client.get("/api/identities").json()
        created_identity = next(item for item in identities if item["id"] == identity_id)

        with (
            patch("app.api.identities.test_smtp_connection", AsyncMock(return_value=(True, "SMTP 连接测试成功"))),
            patch("app.api.identities.test_imap_connection", AsyncMock(return_value=(True, "IMAP 连接测试成功"))),
            patch(
                "app.api.llm_profiles.probe_llm_profile",
                AsyncMock(
                    return_value=self._build_probe_result(
                        ok=True,
                        message="模型连通性测试成功",
                        resolved_base_url="https://api.example.com/v1",
                        response_preview="READY",
                    ),
                ),
            ),
        ):
            smtp_result = self.client.post(f"/api/identities/{identity_id}/smtp-test")
            imap_result = self.client.post(f"/api/identities/{identity_id}/imap-test")
            llm_result = self.client.post(f"/api/llm-profiles/{llm_id}/test")

        self.assertEqual(smtp_result.status_code, 200)
        self.assertTrue(smtp_result.json()["ok"])
        self.assertEqual(imap_result.status_code, 200)
        self.assertTrue(imap_result.json()["ok"])
        self.assertEqual(llm_result.status_code, 200)
        self.assertTrue(llm_result.json()["ok"])
        self.assertEqual(created_identity["smtp_username"], "sender@example.com")
        self.assertEqual(created_identity["imap_host"], "imap.example.com")
        self.assertEqual(created_identity["imap_port"], 993)
        self.assertEqual(created_identity["imap_username"], "sender@example.com")
        self.assertEqual(created_identity["imap_password"], "secret")

    def test_identity_accepts_profile_name_and_sender_name_with_name_compatibility(self) -> None:
        payload = self._build_identity_payload(
            with_imap=False,
            outreach_template_subject="申请与{{name}}老师交流",
            outreach_template_body_text="老师您好，我是{{sender_name}}。",
        )
        payload["name"] = "兼容配置名称"
        payload["profile_name"] = "博士申请配置"
        payload["sender_name"] = "王同学"
        payload["email_address"] = "sender-profile-name@example.com"
        payload["smtp_username"] = "sender-profile-name@example.com"

        response = self.client.post("/api/identities", json=payload)

        self.assertEqual(response.status_code, 201, msg=response.text)
        body = response.json()
        self.assertEqual(body["name"], "博士申请配置")
        self.assertEqual(body["profile_name"], "博士申请配置")
        self.assertEqual(body["sender_name"], "王同学")

        list_payload = self.client.get("/api/identities").json()
        created = next(item for item in list_payload if item["id"] == body["id"])
        self.assertEqual(created["name"], "博士申请配置")
        self.assertEqual(created["profile_name"], "博士申请配置")
        self.assertEqual(created["sender_name"], "王同学")

    def test_identity_legacy_name_populates_profile_and_sender_name(self) -> None:
        payload = self._build_identity_payload(
            with_imap=False,
            outreach_template_subject="申请与{{name}}老师交流",
            outreach_template_body_text="老师您好，我是{{sender_name}}。",
        )
        payload["email_address"] = "legacy-name@example.com"
        payload["smtp_username"] = "legacy-name@example.com"
        payload.pop("profile_name", None)
        payload.pop("sender_name", None)
        payload["name"] = "旧身份名称"

        response = self.client.post("/api/identities", json=payload)

        self.assertEqual(response.status_code, 201, msg=response.text)
        body = response.json()
        self.assertEqual(body["name"], "旧身份名称")
        self.assertEqual(body["profile_name"], "旧身份名称")
        self.assertEqual(body["sender_name"], "旧身份名称")

    def test_system_settings_endpoint_is_removed(self) -> None:
        response = self.client.get("/api/system-settings")

        self.assertEqual(response.status_code, 404)

    def test_llm_model_catalog_endpoint(self) -> None:
        llm_id = self._create_llm()

        with patch(
            "app.api.llm_profiles.fetch_llm_profile_models",
            AsyncMock(
                return_value=self._build_model_catalog_result(
                    ok=True,
                    message="已获取 2 个模型",
                    resolved_base_url="https://api.example.com/v1",
                    models=["gpt-5.4", "gpt-5.4-mini"],
                    selected_model_available=True,
                ),
            ),
        ):
            response = self.client.get(f"/api/llm-profiles/{llm_id}/models")

        self.assertEqual(response.status_code, 200)
        body = response.json()
        self.assertTrue(body["ok"])
        self.assertEqual(body["models"], ["gpt-5.4", "gpt-5.4-mini"])
        self.assertTrue(body["selected_model_available"])

    def test_identity_template_import_endpoint_supports_unsaved_identity_flow(self) -> None:
        response = self.client.post(
            "/api/identities/template-import",
            files={
                "file": (
                    "template.html",
                    io.BytesIO("<p>{{name}}老师您好，</p><p>我是{{sender_name}}。</p>".encode("utf-8")),
                    "text/html",
                )
            },
        )

        self.assertEqual(response.status_code, 200, msg=response.text)
        payload = response.json()
        self.assertIsNone(payload["subject"])
        self.assertEqual(payload["format_name"], "html")
        self.assertEqual(payload["body_text"], "{{name}}老师您好，\n\n我是{{sender_name}}。")
        self.assertIn("<p>{{name}}老师您好，</p>", payload["body_html"])

    def test_identity_template_import_endpoint_requires_file_name(self) -> None:
        boundary = "X-BOUNDARY"
        body = (
            f"--{boundary}\r\n"
            'Content-Disposition: form-data; name="file"; filename=""\r\n'
            "Content-Type: text/plain\r\n\r\n"
            "老师您好\r\n"
            f"--{boundary}--\r\n"
        ).encode("utf-8")
        response = self.client.post(
            "/api/identities/template-import",
            content=body,
            headers={"Content-Type": f"multipart/form-data; boundary={boundary}"},
        )

        self.assertEqual(response.status_code, 400, msg=response.text)
        self.assertEqual(response.json()["detail"], "请选择模板文件")

    def test_identity_allows_missing_template_subject_in_all_modes(self) -> None:
        for mode in ("llm", "template"):
            with self.subTest(mode=mode):
                payload = self._build_identity_payload(
                    with_imap=False,
                    outreach_generation_mode=mode,
                    outreach_template_subject=None,
                    outreach_template_body_text="老师您好，我是{{sender_name}}。",
                    outreach_template_body_html="<p>老师您好，我是{{sender_name}}。</p>",
                )
                payload["email_address"] = f"sender-{mode}@example.com"
                payload["smtp_username"] = f"sender-{mode}@example.com"
                response = self.client.post(
                    "/api/identities",
                    json=payload,
                )

                self.assertEqual(response.status_code, 201, msg=response.text)
                self.assertIsNone(response.json()["outreach_template_subject"])

    def test_identity_allows_missing_plain_text_template_body_even_when_html_exists(self) -> None:
        response = self.client.post(
            "/api/identities",
            json=self._build_identity_payload(
                with_imap=False,
                outreach_generation_mode="llm",
                outreach_template_subject="申请与{{name}}老师交流",
                outreach_template_body_text=None,
                outreach_template_body_html="<p>老师您好，我是{{sender_name}}。</p>",
            ),
        )

        self.assertEqual(response.status_code, 201, msg=response.text)
        self.assertIsNone(response.json()["outreach_template_body_text"])
        self.assertEqual(response.json()["outreach_template_body_html"], "<p>老师您好，我是{{sender_name}}。</p>")

    def test_identity_update_allows_incomplete_template_defaults(self) -> None:
        cases = [
            {
                "name": "只缺主题",
                "subject": None,
                "body_text": "老师您好，我是{{sender_name}}。",
                "body_html": "<p>老师您好，我是{{sender_name}}。</p>",
            },
            {
                "name": "只缺纯文本正文",
                "subject": "申请与{{name}}老师交流",
                "body_text": None,
                "body_html": "<p>老师您好，我是{{sender_name}}。</p>",
            },
            {
                "name": "主题和纯文本正文都缺",
                "subject": None,
                "body_text": None,
                "body_html": "<p>老师您好，我是{{sender_name}}。</p>",
            },
        ]

        for case in cases:
            with self.subTest(case=case["name"]):
                unique_email = f"sender-update-{case['name']}@example.com"
                create_payload = self._build_identity_payload(
                    with_imap=False,
                    outreach_generation_mode="llm",
                    outreach_template_subject="申请与{{name}}老师交流",
                    outreach_template_body_text="老师您好，我是{{sender_name}}，关注到您在{{research_direction}}方向的工作。",
                )
                create_payload["email_address"] = unique_email
                create_payload["smtp_username"] = unique_email
                create_response = self.client.post("/api/identities", json=create_payload)
                self.assertEqual(create_response.status_code, 201, msg=create_response.text)
                identity_id = create_response.json()["id"]
                update_payload = self._build_identity_payload(
                    with_imap=False,
                    outreach_generation_mode="llm",
                    outreach_template_subject=case["subject"],
                    outreach_template_body_text=case["body_text"],
                    outreach_template_body_html=case["body_html"],
                )
                update_payload["email_address"] = unique_email
                update_payload["smtp_username"] = unique_email
                response = self.client.put(
                    f"/api/identities/{identity_id}",
                    json=update_payload,
                )

                self.assertEqual(response.status_code, 200, msg=response.text)
                self.assertEqual(response.json()["outreach_template_subject"], case["subject"])
                self.assertEqual(response.json()["outreach_template_body_text"], case["body_text"])
                self.assertEqual(response.json()["outreach_template_body_html"], case["body_html"])

    def test_llm_structured_result_validation_rejects_invalid_json(self) -> None:
        from app.services.llm_runtime import DraftGenerationResult, LLMRuntimeError, parse_structured_result

        with self.assertRaises(LLMRuntimeError):
            parse_structured_result('{"subject":"only-subject"}', DraftGenerationResult)

    def test_professor_management_crud_archive_restore_and_dashboard_filtering(self) -> None:
        create_response = self.client.post(
            "/api/professors",
            json={
                "name": "张教授",
                "email": "zhang@example.edu",
                "title": "教授",
                "university": "Example University",
                "school": "School of AI",
                "department": "Computer Science",
                "research_direction": "Large language models",
                "recent_papers": ["Paper A", "Paper B"],
                "profile_url": "https://example.edu/zhang",
                "source_url": "https://example.edu/faculty",
            },
        )
        self.assertEqual(create_response.status_code, 201, msg=create_response.text)
        professor_id = create_response.json()["id"]

        update_response = self.client.patch(
            f"/api/professors/{professor_id}",
            json={
                "name": "张教授",
                "email": "zhang@example.edu",
                "title": "副教授",
                "university": "Example University",
                "school": "School of AI",
                "department": "Computer Science",
                "research_direction": "Agent systems",
                "recent_papers": ["Paper C"],
                "profile_url": "https://example.edu/zhang-new",
                "source_url": "https://example.edu/faculty",
            },
        )
        self.assertEqual(update_response.status_code, 200)
        self.assertEqual(update_response.json()["title"], "副教授")
        self.assertEqual(update_response.json()["recent_papers"], ["Paper C"])

        active_list = self.client.get("/api/professors/management", params={"archived": "active"})
        self.assertEqual(active_list.status_code, 200)
        self.assertEqual(len(active_list.json()), 1)

        archive_response = self.client.post(f"/api/professors/{professor_id}/archive")
        self.assertEqual(archive_response.status_code, 200)
        self.assertEqual(archive_response.json()["affected_count"], 1)

        dashboard_list = self.client.get("/api/professors")
        self.assertEqual(dashboard_list.status_code, 200)
        self.assertEqual(dashboard_list.json(), [])

        archived_list = self.client.get("/api/professors/management", params={"archived": "archived"})
        self.assertEqual(archived_list.status_code, 200)
        self.assertEqual(len(archived_list.json()), 1)
        self.assertIsNotNone(archived_list.json()[0]["archived_at"])

        restore_response = self.client.post(f"/api/professors/{professor_id}/restore")
        self.assertEqual(restore_response.status_code, 200)
        self.assertEqual(restore_response.json()["affected_count"], 1)

        second_professor = self.client.post(
            "/api/professors",
            json={
                "name": "王教授",
                "email": "wang-prof@example.edu",
                "title": "Professor",
                "university": "Example University",
                "school": "School of AI",
                "department": "Computer Science",
                "research_direction": "Information extraction",
                "recent_papers": [],
                "profile_url": None,
                "source_url": None,
            },
        ).json()
        bulk_archive_response = self.client.post(
            "/api/professors/bulk-archive",
            json={"ids": [professor_id, second_professor["id"]]},
        )
        self.assertEqual(bulk_archive_response.status_code, 200)
        self.assertEqual(bulk_archive_response.json()["affected_count"], 2)

        restored_after_bulk = self.client.get("/api/professors").json()
        self.assertEqual(restored_after_bulk, [])

        self.client.post(f"/api/professors/{professor_id}/restore")

        restored_dashboard = self.client.get("/api/professors")
        self.assertEqual(restored_dashboard.status_code, 200)
        self.assertEqual(len(restored_dashboard.json()), 1)
        self.assertEqual(restored_dashboard.json()[0]["name"], "张教授")

    def test_professor_dashboard_returns_contact_state_labels(self) -> None:
        identity_id = self._create_identity(with_imap=False)
        llm_id = self._create_llm()
        professor_cases = [
            ("未联系导师", "dashboard-not-contacted@example.edu", None, "not_contacted"),
            ("准备中导师", "dashboard-preparing@example.edu", "matched", "preparing"),
            ("approved 导师", "dashboard-approved@example.edu", "approved", "ready_to_send"),
            ("待发送导师", "dashboard-ready@example.edu", "scheduled", "ready_to_send"),
            ("已联系导师", "dashboard-contacted@example.edu", "sent", "contacted"),
            ("已回复导师", "dashboard-replied@example.edu", "reply_detected", "replied"),
            ("send_failed 导师", "dashboard-send-failed@example.edu", "send_failed", "needs_attention"),
            ("需处理导师", "dashboard-needs-attention@example.edu", "canceled", "needs_attention"),
        ]

        professor_ids: dict[str, int] = {}
        task_ids: dict[str, int] = {}

        for name, email, task_status, _expected_status in professor_cases:
            create_response = self.client.post(
                "/api/professors",
                json={
                    "name": name,
                    "email": email,
                    "title": "Professor",
                    "university": "Example University",
                    "school": "School of AI",
                    "department": "Computer Science",
                    "research_direction": "Large language models",
                    "recent_papers": [],
                    "profile_url": None,
                    "source_url": None,
                },
            )
            self.assertEqual(create_response.status_code, 201, msg=create_response.text)
            professor_id = create_response.json()["id"]
            professor_ids[email] = professor_id

            if task_status is None:
                continue

            ensure_response = self.client.post(
                f"/api/workspaces/{professor_id}/ensure-task",
                params={"identity_id": identity_id, "llm_profile_id": llm_id},
            )
            self.assertEqual(ensure_response.status_code, 200, msg=ensure_response.text)
            task_ids[email] = ensure_response.json()["current_task"]["id"]

        connection = sqlite3.connect(self.db_path)
        try:
            for _name, email, task_status, _expected_status in professor_cases:
                task_id = task_ids.get(email)
                if task_id is None or task_status is None:
                    continue
                connection.execute(
                    """
                    UPDATE email_tasks
                    SET status = ?, cancellation_reason = ?
                    WHERE id = ?
                    """,
                    (
                        task_status,
                        "batch_stopped" if task_status == "canceled" else None,
                        task_id,
                    ),
                )
            connection.commit()
        finally:
            connection.close()

        response = self.client.get(
            "/api/professors",
            params={"identity_id": identity_id, "llm_profile_id": llm_id},
        )

        self.assertEqual(response.status_code, 200, msg=response.text)
        payload_by_id = {item["id"]: item for item in response.json()}
        for _name, email, _task_status, expected_status in professor_cases:
            payload = payload_by_id[professor_ids[email]]
            self.assertEqual(payload["status"], expected_status)
            self.assertNotIn(
                payload["status"],
                {"matched", "scheduled", "sent", "skipped", "send_failed"},
            )

    def test_professor_dashboard_prioritizes_existing_contact_over_follow_up_draft(self) -> None:
        identity_id = self._create_identity(with_imap=False)
        llm_id = self._create_llm()
        professor_response = self.client.post(
            "/api/professors",
            json={
                "name": "已联系后跟进导师",
                "email": "contacted-follow-up@example.edu",
                "title": "Professor",
                "university": "Example University",
                "school": "School of AI",
                "department": "Computer Science",
                "research_direction": "Large language models",
                "recent_papers": [],
                "profile_url": None,
                "source_url": None,
            },
        )
        self.assertEqual(professor_response.status_code, 201, msg=professor_response.text)
        professor_id = professor_response.json()["id"]

        ensure_response = self.client.post(
            f"/api/workspaces/{professor_id}/ensure-task",
            params={"identity_id": identity_id, "llm_profile_id": llm_id},
        )
        self.assertEqual(ensure_response.status_code, 200, msg=ensure_response.text)
        parent_task_id = ensure_response.json()["current_task"]["id"]

        connection = sqlite3.connect(self.db_path)
        try:
            connection.execute(
                """
                UPDATE email_tasks
                SET status = 'sent', sent_at = CURRENT_TIMESTAMP
                WHERE id = ?
                """,
                (parent_task_id,),
            )
            connection.execute(
                """
                INSERT INTO email_logs (
                    email_task_id, identity_id, llm_profile_id, professor_id,
                    direction, subject, content, rfc_message_id
                )
                VALUES (?, ?, ?, ?, 'sent', 'hello', 'hello body', '<sent@example.edu>')
                """,
                (parent_task_id, identity_id, llm_id, professor_id),
            )
            connection.execute(
                """
                INSERT INTO email_tasks (
                    source, parent_task_id, identity_id, llm_profile_id,
                    professor_id, status, created_at, updated_at
                )
                VALUES ('manual', ?, ?, ?, ?, 'matched', datetime('now', '+1 minute'), datetime('now', '+1 minute'))
                """,
                (parent_task_id, identity_id, llm_id, professor_id),
            )
            connection.commit()
        finally:
            connection.close()

        response = self.client.get(
            "/api/professors",
            params={"identity_id": identity_id, "llm_profile_id": llm_id},
        )

        self.assertEqual(response.status_code, 200, msg=response.text)
        professor = next(item for item in response.json() if item["id"] == professor_id)
        self.assertEqual(professor["status"], "contacted")

        connection = sqlite3.connect(self.db_path)
        try:
            connection.execute(
                "UPDATE email_tasks SET status = 'reply_detected', is_replied = 1 WHERE id = ?",
                (parent_task_id,),
            )
            connection.commit()
        finally:
            connection.close()

        replied_response = self.client.get(
            "/api/professors",
            params={"identity_id": identity_id, "llm_profile_id": llm_id},
        )

        self.assertEqual(replied_response.status_code, 200, msg=replied_response.text)
        replied_professor = next(item for item in replied_response.json() if item["id"] == professor_id)
        self.assertEqual(replied_professor["status"], "replied")

    def test_professor_template_download_and_import_file_upserts_existing_records(self) -> None:
        csv_template = self.client.get("/api/professors/template", params={"format": "csv"})
        xlsx_template = self.client.get("/api/professors/template", params={"format": "xlsx"})
        self.assertEqual(csv_template.status_code, 200)
        self.assertIn("professors_import_template.csv", csv_template.headers["content-disposition"])
        self.assertIn("# 导师导入模板", csv_template.text)
        self.assertIn("# name：导师姓名，必填。示例：张明远", csv_template.text)
        self.assertIn("# title：导师职称。示例：教授", csv_template.text)
        self.assertIn("# university：学校名称。示例：示例大学", csv_template.text)
        self.assertIn("# school：学院名称。示例：人工智能学院", csv_template.text)
        self.assertIn("# department：院系或系所。示例：计算机科学系", csv_template.text)
        self.assertIn("# research_direction：研究方向，多个方向用中文分号 ； 分隔。示例：大语言模型；智能体；信息抽取", csv_template.text)
        self.assertIn("# recent_papers：近期论文，多篇用 | 分隔；最多保留前 8 篇。示例：Paper A|Paper B", csv_template.text)
        self.assertIn("name,email,title", csv_template.text)
        self.assertIn("示例：张明远,zhang@example.edu,教授,示例大学,人工智能学院,计算机科学系,大语言模型；智能体；信息抽取", csv_template.text)
        self.assertEqual(xlsx_template.status_code, 200)
        self.assertIn("professors_import_template.xlsx", xlsx_template.headers["content-disposition"])
        workbook_from_template = load_workbook(io.BytesIO(xlsx_template.content))
        template_sheet = workbook_from_template.active
        template_values = list(template_sheet.iter_rows(values_only=True))
        self.assertEqual(template_values[0][0], "# 导师导入模板")
        self.assertEqual(template_values[3][0], "# name：导师姓名，必填。示例：张明远")
        self.assertEqual(template_values[5][0], "# title：导师职称。示例：教授")
        self.assertEqual(template_values[6][0], "# university：学校名称。示例：示例大学")
        template_headers = list(template_values[13])
        self.assertEqual(
            template_headers,
            [
                "name",
                "email",
                "title",
                "university",
                "school",
                "department",
                "research_direction",
                "recent_papers",
                "profile_url",
                "source_url",
            ],
        )
        self.assertEqual(template_values[14][0], "示例：张明远")
        self.assertEqual(
            list(template_values[14][2:7]),
            ["教授", "示例大学", "人工智能学院", "计算机科学系", "大语言模型；智能体；信息抽取"],
        )

        created_professor = self.client.post(
            "/api/professors",
            json={
                "name": "李教授",
                "email": "li@example.edu",
                "title": "Professor",
                "university": "Legacy University",
                "school": "School of Computing",
                "department": "CS",
                "research_direction": "Legacy direction",
                "recent_papers": ["Legacy Paper"],
                "profile_url": None,
                "source_url": None,
            },
        ).json()
        professor_id = created_professor["id"]
        self.client.post(f"/api/professors/{professor_id}/archive")

        csv_content = (
            "# 导师导入模板\n"
            "# 从字段名下一行开始填写；说明行和示例行可以保留，系统导入时会自动忽略\n"
            "# 必填字段：name, email\n"
            "name,email,title,university,school,department,research_direction,recent_papers,profile_url,source_url\n"
            "示例：张明远,zhang@example.edu,教授,示例大学,人工智能学院,计算机科学系,大语言模型；智能体；信息抽取,Paper A|Paper B,https://example.edu/zhang,https://example.edu/faculty\n"
            "李教授,li@example.edu,副教授,New University,School of AI,AI,Updated direction,Paper 1|Paper 2|Paper 3|Paper 4|Paper 5|Paper 6|Paper 7|Paper 8|Paper 9|Paper 10,https://example.edu/li,https://example.edu/faculty\n"
            "王老师,wang@example.edu,Assistant Professor,Another University,School,Dept,Direction,Paper 3,,\n"
            "坏数据,not-an-email,Professor,Bad University,School,Dept,Direction,Paper X,,\n"
        ).encode("utf-8-sig")
        csv_import = self.client.post(
            "/api/professors/import-file",
            files={"file": ("professors.csv", io.BytesIO(csv_content), "text/csv")},
        )
        self.assertEqual(csv_import.status_code, 200, msg=csv_import.text)
        csv_body = csv_import.json()
        self.assertEqual(csv_body["inserted_count"], 1)
        self.assertEqual(csv_body["updated_count"], 1)
        self.assertEqual(csv_body["failed_count"], 1)

        refreshed = self.client.get("/api/professors/management", params={"archived": "active"}).json()
        li_professor = next(item for item in refreshed if item["email"] == "li@example.edu")
        self.assertEqual(li_professor["title"], "副教授")
        self.assertEqual(
            li_professor["recent_papers"],
            ["Paper 1", "Paper 2", "Paper 3", "Paper 4", "Paper 5", "Paper 6", "Paper 7", "Paper 8"],
        )
        self.assertIsNone(li_professor["archived_at"])

        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["# 导师导入模板"])
        sheet.append(["# 从字段名下一行开始填写；说明行和示例行可以保留，系统导入时会自动忽略"])
        sheet.append(["# 必填字段：name, email"])
        sheet.append(
            [
                "name",
                "email",
                "title",
                "university",
                "school",
                "department",
                "research_direction",
                "recent_papers",
                "profile_url",
                "source_url",
            ]
        )
        sheet.append(
            [
                "示例：张明远",
                "zhang@example.edu",
                "教授",
                "示例大学",
                "人工智能学院",
                "计算机科学系",
                "大语言模型；智能体；信息抽取",
                "Paper A|Paper B",
                "https://example.edu/zhang",
                "https://example.edu/faculty",
            ]
        )
        sheet.append(
            [
                "王老师",
                "wang@example.edu",
                "Professor",
                "Updated University",
                "New School",
                "New Dept",
                "Updated research",
                "Paper 8|Paper 9",
                "https://example.edu/wang",
                "https://example.edu/source",
            ]
        )
        buffer = io.BytesIO()
        workbook.save(buffer)

        xlsx_import = self.client.post(
            "/api/professors/import-file",
            files={
                "file": (
                    "professors.xlsx",
                    io.BytesIO(buffer.getvalue()),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        self.assertEqual(xlsx_import.status_code, 200, msg=xlsx_import.text)
        self.assertEqual(xlsx_import.json()["inserted_count"], 0)
        self.assertEqual(xlsx_import.json()["updated_count"], 1)

        management_all = self.client.get("/api/professors/management", params={"archived": "all"}).json()
        wang_professor = next(item for item in management_all if item["email"] == "wang@example.edu")
        self.assertEqual(wang_professor["university"], "Updated University")
        self.assertEqual(wang_professor["recent_papers"], ["Paper 8", "Paper 9"])

    def test_workspace_endpoint_without_existing_task_returns_empty_thread(self) -> None:
        identity_id = self._create_identity(with_imap=False)
        llm_id = self._create_llm()

        professor_response = self.client.post(
            "/api/professors",
            json={
                "name": "空白导师",
                "email": "blank@example.edu",
                "title": "Professor",
                "university": "Example University",
                "school": "School of Computing",
                "department": "Computer Science",
                "research_direction": "Distributed systems",
                "recent_papers": [],
                "profile_url": None,
                "source_url": None,
            },
        )
        self.assertEqual(professor_response.status_code, 201, msg=professor_response.text)
        professor_id = professor_response.json()["id"]

        workspace_response = self.client.get(
            f"/api/workspaces/{professor_id}",
            params={"identity_id": identity_id, "llm_profile_id": llm_id},
        )
        self.assertEqual(workspace_response.status_code, 200, msg=workspace_response.text)

        payload = workspace_response.json()
        self.assertEqual(payload["professor"]["id"], professor_id)
        self.assertIsNone(payload["current_task"]["id"])
        self.assertEqual(payload["current_task"]["fit_points"], [])
        self.assertEqual(payload["current_task"]["risk_points"], [])
        self.assertEqual(payload["current_task"]["match_keywords"], [])
        self.assertEqual(payload["messages"], [])

    def test_workspace_ensure_task_creates_and_reuses_personal_task(self) -> None:
        identity_id = self._create_identity(with_imap=False)
        llm_id = self._create_llm()
        material_id = self._upload_material(
            identity_id,
            filename="resume.txt",
            content=b"My research focuses on agents and information extraction.",
            material_type="resume",
        )

        professor_response = self.client.post(
            "/api/professors",
            json={
                "name": "可直达工作区导师",
                "email": "direct-workspace@example.edu",
                "title": "Professor",
                "university": "Example University",
                "school": "School of Computing",
                "department": "Computer Science",
                "research_direction": "Agent systems",
                "recent_papers": [],
                "profile_url": None,
                "source_url": None,
            },
        )
        self.assertEqual(professor_response.status_code, 201, msg=professor_response.text)
        professor_id = professor_response.json()["id"]

        first_response = self.client.post(
            f"/api/workspaces/{professor_id}/ensure-task",
            params={"identity_id": identity_id, "llm_profile_id": llm_id},
        )
        self.assertEqual(first_response.status_code, 200, msg=first_response.text)
        first_payload = first_response.json()
        self.assertIsNotNone(first_payload["current_task"]["id"])
        self.assertEqual(first_payload["current_task"]["batch_task_id"], None)
        self.assertEqual(first_payload["current_task"]["status"], "discovered")
        self.assertEqual(first_payload["current_task"]["primary_material_id"], material_id)

        second_response = self.client.post(
            f"/api/workspaces/{professor_id}/ensure-task",
            params={"identity_id": identity_id, "llm_profile_id": llm_id},
        )
        self.assertEqual(second_response.status_code, 200, msg=second_response.text)
        second_payload = second_response.json()
        self.assertEqual(
            second_payload["current_task"]["id"],
            first_payload["current_task"]["id"],
        )
        self.assertEqual(second_payload["messages"], [])

    def test_template_mode_can_generate_draft_without_primary_material(self) -> None:
        identity_id = self._create_identity(with_imap=False)
        llm_id = self._create_llm()

        update_response = self.client.put(
            f"/api/identities/{identity_id}",
            json={
                "name": "模板身份",
                "email_address": "sender@example.com",
                "smtp_host": "smtp.example.com",
                "smtp_port": 465,
                "smtp_username": "sender@example.com",
                "smtp_password": "secret",
                "imap_host": "imap.example.com",
                "imap_port": 993,
                "imap_username": "sender@example.com",
                "imap_password": "secret",
                "default_language": "zh-CN",
                "outreach_generation_mode": "template",
                "outreach_template_subject": "申请与{{name}}老师交流",
                "outreach_template_body_text": "{{name}}老师您好，我是{{sender_name}}。",
                "outreach_template_body_html": "<p>{{name}}老师您好，我是{{sender_name}}。</p>",
                "match_threshold": None,
                "daily_send_limit": None,
                "send_interval_min": None,
                "send_interval_max": None,
                "same_domain_cooldown_minutes": None,
                "is_default": True,
            },
        )
        self.assertEqual(update_response.status_code, 200, msg=update_response.text)

        professor_response = self.client.post(
            "/api/professors",
            json={
                "name": "模板导师",
                "email": "template@example.edu",
                "title": "Professor",
                "university": "Example University",
                "school": "School of Computing",
                "department": "Computer Science",
                "research_direction": "Agents",
                "recent_papers": [],
                "profile_url": None,
                "source_url": None,
            },
        )
        self.assertEqual(professor_response.status_code, 201, msg=professor_response.text)
        professor_id = professor_response.json()["id"]

        ensure_response = self.client.post(
            f"/api/workspaces/{professor_id}/ensure-task",
            params={"identity_id": identity_id, "llm_profile_id": llm_id},
        )
        self.assertEqual(ensure_response.status_code, 200, msg=ensure_response.text)
        task_id = ensure_response.json()["current_task"]["id"]

        generate_response = self.client.post(
            f"/api/email-tasks/{task_id}/generate-draft",
        )
        self.assertEqual(generate_response.status_code, 200, msg=generate_response.text)
        payload = generate_response.json()
        self.assertEqual(payload["current_task"]["status"], "review_required")
        self.assertEqual(payload["current_task"]["generated_subject"], "申请与模板导师老师交流")
        self.assertIn("模板导师老师您好", payload["current_task"]["generated_content_text"])

    def test_manual_send_renders_subject_and_body_placeholders(self) -> None:
        identity_id = self._create_identity(with_imap=False)
        llm_id = self._create_llm()
        professor_response = self.client.post(
            "/api/professors",
            json={
                "name": "主题导师",
                "email": "subject@example.edu",
                "title": "Professor",
                "university": "Example University",
                "school": "School of Computing",
                "department": "Computer Science",
                "research_direction": "Agents",
                "recent_papers": [],
                "profile_url": None,
                "source_url": None,
            },
        )
        self.assertEqual(professor_response.status_code, 201, msg=professor_response.text)
        professor_id = professor_response.json()["id"]
        ensure_response = self.client.post(
            f"/api/workspaces/{professor_id}/ensure-task",
            params={"identity_id": identity_id, "llm_profile_id": llm_id},
        )
        self.assertEqual(ensure_response.status_code, 200, msg=ensure_response.text)
        task_id = ensure_response.json()["current_task"]["id"]

        with patch(
            "app.services.task_runtime.mail_runtime.send_email",
            AsyncMock(return_value=self._build_send_result(message_id="<subject-render@example.com>", provider_payload={})),
        ) as mocked_send:
            response = self.client.post(
                f"/api/email-tasks/{task_id}/approve-and-send",
                json={
                    "subject": "申请与{{name}}老师交流",
                    "body_text": "{{name}}老师您好，我是{{sender_name}}。",
                    "body_html": "<p>{{name}}老师您好，我是{{sender_name}}。</p>",
                    "selected_material_ids": [],
                },
            )

        self.assertEqual(response.status_code, 200, msg=response.text)
        kwargs = mocked_send.await_args.kwargs
        self.assertEqual(kwargs["subject"], "申请与主题导师老师交流")
        self.assertIn("主题导师老师您好", kwargs["body_text"])
        self.assertNotIn("{{name}}", kwargs["body_html"])
        self.assertEqual(response.json()["current_task"]["approved_subject"], "申请与{{name}}老师交流")

    def test_identity_llm_mode_allows_empty_template_defaults(self) -> None:
        response = self.client.post(
            "/api/identities",
            json=self._build_identity_payload(
                with_imap=False,
                outreach_generation_mode="llm",
                outreach_template_subject=None,
                outreach_template_body_text=None,
                outreach_template_body_html=None,
            ),
        )

        self.assertEqual(response.status_code, 201, msg=response.text)
        self.assertIsNone(response.json()["outreach_template_subject"])
        self.assertIsNone(response.json()["outreach_template_body_text"])
        self.assertIsNone(response.json()["outreach_template_body_html"])

    def test_material_upload_open_download_set_primary_and_delete(self) -> None:
        identity_id = self._create_identity(with_imap=False)

        resume_material_id = self._upload_material(
            identity_id,
            filename="cv.txt",
            content=b"My research background is in information extraction.",
            material_type="resume",
        )
        image_material_id = self._upload_material(
            identity_id,
            filename="poster.png",
            content=b"fake-image",
            material_type="portfolio",
        )

        identities = self.client.get("/api/identities").json()
        identity = next(item for item in identities if item["id"] == identity_id)
        self.assertEqual(identity["current_primary_material_id"], resume_material_id)
        self.assertEqual(len(identity["materials"]), 2)

        invalid_primary_response = self.client.post(f"/api/materials/{image_material_id}/set-primary")
        self.assertEqual(invalid_primary_response.status_code, 400)

        open_response = self.client.get(f"/api/materials/{resume_material_id}/open")
        self.assertEqual(open_response.status_code, 200)
        self.assertIn("inline", open_response.headers.get("content-disposition", ""))

        download_response = self.client.get(f"/api/materials/{resume_material_id}/download")
        self.assertEqual(download_response.status_code, 200)
        self.assertIn("cv.txt", download_response.headers.get("content-disposition", ""))

        delete_primary_response = self.client.delete(f"/api/materials/{resume_material_id}")
        self.assertEqual(delete_primary_response.status_code, 204)

        identity_after_primary_delete = next(
            item for item in self.client.get("/api/identities").json() if item["id"] == identity_id
        )
        self.assertIsNone(identity_after_primary_delete["current_primary_material_id"])
        self.assertEqual(len(identity_after_primary_delete["materials"]), 1)

        delete_response = self.client.delete(f"/api/materials/{image_material_id}")
        self.assertEqual(delete_response.status_code, 204)

        refreshed_identity = next(
            item for item in self.client.get("/api/identities").json() if item["id"] == identity_id
        )
        self.assertEqual(len(refreshed_identity["materials"]), 0)

    def test_material_upload_keeps_working_when_text_extraction_fails(self) -> None:
        identity_id = self._create_identity(with_imap=False)

        with patch("app.services.file_storage._get_markitdown") as mocked_markitdown:
            mocked_markitdown.return_value.convert.side_effect = RuntimeError("boom")
            response = self.client.post(
                f"/api/identities/{identity_id}/materials",
                files={"file": ("transcript.pdf", b"%PDF-pretend-transcript", "application/pdf")},
                data={"material_type": "transcript"},
            )

        self.assertEqual(response.status_code, 201, msg=response.text)
        body = response.json()
        self.assertEqual(body["material_type"], "transcript")
        self.assertEqual(body["display_name"], "transcript")

    def test_primary_material_text_is_extracted_on_demand_when_generating_draft(self) -> None:
        identity_id = self._create_identity(with_imap=False)
        llm_id = self._create_llm()
        material_id = self._upload_material(
            identity_id,
            filename="resume.txt",
            content=b"My research focuses on information extraction and agents.",
            material_type="resume",
        )

        connection = sqlite3.connect(self.db_path)
        try:
            row = connection.execute(
                "SELECT extracted_text FROM identity_materials WHERE id = ?",
                (material_id,),
            ).fetchone()
        finally:
            connection.close()
        self.assertEqual(row[0], None)

        self.client.post("/api/professors/import-sample")
        professor_id = self.client.get("/api/professors").json()[0]["id"]
        task_response = self.client.post(
            "/api/batch-tasks",
            json={
                "identity_id": identity_id,
                "llm_profile_id": llm_id,
                "name": "按需解析默认材料",
                "professor_ids": [professor_id],
                "schedule_type": "immediate",
                "window_start_time": None,
                "window_end_time": None,
                "emails_per_window": None,
                "primary_material_id": material_id,
                "email_subject": "申请与{{name}}老师交流",
                "email_body": "老师您好，我是{{sender_name}}，关注到您在{{research_direction}}方向的工作。",
                "selected_material_ids": None,
            },
        )
        self.assertEqual(task_response.status_code, 201)

        workspace = self.client.get(
            f"/api/workspaces/{professor_id}",
            params={"identity_id": identity_id, "llm_profile_id": llm_id},
        )
        task_id = workspace.json()["current_task"]["id"]

        with patch(
            "app.services.task_runtime.llm_runtime.generate_draft_content",
            AsyncMock(
                return_value=self._build_draft_generation_result(
                    subject="测试草稿",
                    body_text="测试正文",
                    body_html="<p>测试正文</p>",
                    prompt_tokens=80,
                    completion_tokens=20,
                    cached_tokens=32,
                ),
            ),
        ):
            regenerate_response = self.client.post(
                f"/api/email-tasks/{task_id}/generate-draft",
            )

        self.assertEqual(regenerate_response.status_code, 200)
        connection = sqlite3.connect(self.db_path)
        try:
            refreshed_row = connection.execute(
                "SELECT extracted_text FROM identity_materials WHERE id = ?",
                (material_id,),
            ).fetchone()
        finally:
            connection.close()
        self.assertIn("information extraction", refreshed_row[0])
        provider_payload = self._latest_email_log_provider_payload()
        self.assertEqual(provider_payload["usage"]["cached_tokens"], 32)

    def test_app_startup_auto_upgrades_stale_database(self) -> None:
        stale_dir = tempfile.TemporaryDirectory()
        stale_db_path = Path(stale_dir.name) / "stale.db"
        stale_env = os.environ.copy()
        stale_env["DATABASE_URL"] = f"sqlite+aiosqlite:///{stale_db_path.as_posix()}"
        stale_env["ENABLE_BACKGROUND_WORKERS"] = "0"

        result = subprocess.run(
            [sys.executable, "-m", "alembic", "upgrade", "c52f8b7d1f43"],
            cwd=BACKEND_DIR,
            env=stale_env,
            capture_output=True,
            text=True,
            check=False,
        )
        self.assertEqual(result.returncode, 0, msg=result.stderr)

        from app.core.config import get_settings
        from app.core.database import dispose_engine, get_engine, get_session_factory
        from main import create_app

        if get_engine.cache_info().currsize:
            asyncio.run(dispose_engine())
        get_session_factory.cache_clear()
        get_settings.cache_clear()
        os.environ["DATABASE_URL"] = stale_env["DATABASE_URL"]

        with TestClient(create_app()) as client:
            response = client.get("/api/ping")
            self.assertEqual(response.status_code, 200)
            self.assertEqual(response.json()["status"], "ok")

        connection = sqlite3.connect(stale_db_path)
        try:
            version = connection.execute(
                "SELECT version_num FROM alembic_version",
            ).fetchone()[0]
        finally:
            connection.close()

        self.assertEqual(version, HEAD_REVISION)

        if get_engine.cache_info().currsize:
            asyncio.run(dispose_engine())
        get_session_factory.cache_clear()
        get_settings.cache_clear()
        stale_dir.cleanup()

    def test_create_scheduled_batch_task_requires_scheduled_dates(self) -> None:
        identity_id = self._create_identity(with_imap=False)
        llm_profile_id = self._create_llm()
        self.client.post("/api/professors/import-sample")
        professor_id = self.client.get("/api/professors").json()[0]["id"]

        response = self.client.post(
            "/api/batch-tasks",
            json={
                "identity_id": identity_id,
                "llm_profile_id": llm_profile_id,
                "name": "定时发送测试",
                "professor_ids": [professor_id],
                "schedule_type": "scheduled",
                "scheduled_dates": [],
                "window_start_time": "09:00",
                "window_end_time": "18:00",
                "emails_per_window": 20,
                "primary_material_id": None,
                "email_subject": "Hello {{导师姓名}}",
                "email_body": "Body",
                "selected_material_ids": None,
                "outreach_generation_mode": "template",
                "outreach_template_subject": "Hello {{导师姓名}}",
                "outreach_template_body_text": "Body",
                "outreach_template_body_html": None,
            },
        )

        self.assertEqual(response.status_code, 400)
        self.assertIn("发送日期", response.json()["detail"])

    def test_batch_task_delete_restore_and_trash_view(self) -> None:
        identity_id = self._create_identity(with_imap=False)
        llm_profile_id = self._create_llm()
        self.client.post("/api/professors/import-sample")
        professor_id = self.client.get("/api/professors").json()[0]["id"]
        created = self.client.post(
            "/api/batch-tasks",
            json={
                "identity_id": identity_id,
                "llm_profile_id": llm_profile_id,
                "name": "可删除批量任务",
                "professor_ids": [professor_id],
                "schedule_type": "immediate",
                "primary_material_id": None,
                "email_subject": "Hello {{导师姓名}}",
                "email_body": "Body",
                "selected_material_ids": None,
                "outreach_generation_mode": "template",
                "outreach_template_subject": "Hello {{导师姓名}}",
                "outreach_template_body_text": "Body",
                "outreach_template_body_html": None,
            },
        )
        self.assertEqual(created.status_code, 201, msg=created.text)
        task_id = created.json()["id"]

        blocked = self.client.post(f"/api/batch-tasks/{task_id}/delete")
        self.assertEqual(blocked.status_code, 400)
        self.assertIn("请先中止/取消任务后再删除", blocked.json()["detail"])

        stopped = self.client.post(f"/api/batch-tasks/{task_id}/stop")
        self.assertEqual(stopped.status_code, 200, msg=stopped.text)
        deleted = self.client.post(f"/api/batch-tasks/{task_id}/delete")
        self.assertEqual(deleted.status_code, 200, msg=deleted.text)
        self.assertIsNotNone(deleted.json()["task"]["deleted_at"])

        repeated_delete = self.client.post(f"/api/batch-tasks/{task_id}/delete")
        self.assertEqual(repeated_delete.status_code, 200, msg=repeated_delete.text)

        current = self.client.get(
            "/api/batch-tasks",
            params={"identity_id": identity_id, "llm_profile_id": llm_profile_id},
        )
        self.assertEqual(current.status_code, 200)
        self.assertEqual(current.json(), [])

        trash = self.client.get(
            "/api/batch-tasks",
            params={
                "identity_id": identity_id,
                "llm_profile_id": llm_profile_id,
                "view": "trash",
            },
        )
        self.assertEqual(trash.status_code, 200)
        self.assertEqual([item["id"] for item in trash.json()], [task_id])

        restored = self.client.post(f"/api/batch-tasks/{task_id}/restore")
        self.assertEqual(restored.status_code, 200, msg=restored.text)
        self.assertIsNone(restored.json()["task"]["deleted_at"])

        repeated_restore = self.client.post(f"/api/batch-tasks/{task_id}/restore")
        self.assertEqual(repeated_restore.status_code, 200, msg=repeated_restore.text)

    def test_create_and_list_match_analysis_jobs(self) -> None:
        identity_id = self._create_identity(with_imap=False)
        llm_id = self._create_llm()
        self._upload_material(
            identity_id,
            filename="resume.txt",
            content=b"AI systems",
            material_type="resume",
        )
        professor_response = self.client.post(
            "/api/professors",
            json={
                "name": "王老师",
                "email": "wang-match@example.edu",
                "title": "Professor",
                "university": "Example University",
                "school": "School of Computing",
                "department": "Computer Science",
                "research_direction": "AI agents",
                "recent_papers": [],
                "profile_url": None,
                "source_url": None,
            },
        )
        self.assertEqual(professor_response.status_code, 201, msg=professor_response.text)
        professor_id = professor_response.json()["id"]

        created = self.client.post(
            "/api/match-analysis-jobs",
            json={
                "identity_id": identity_id,
                "llm_profile_id": llm_id,
                "professor_ids": [professor_id],
            },
        )
        self.assertEqual(created.status_code, 201, msg=created.text)
        self.assertEqual(created.json()["target_count"], 1)

        listed = self.client.get(
            "/api/match-analysis-jobs",
            params={"identity_id": identity_id, "llm_profile_id": llm_id},
        )
        self.assertEqual(listed.status_code, 200)
        self.assertEqual(len(listed.json()), 1)

    def test_match_analysis_job_delete_restore_and_trash_view(self) -> None:
        identity_id = self._create_identity(with_imap=False)
        llm_id = self._create_llm()
        self._upload_material(
            identity_id,
            filename="resume.txt",
            content=b"AI systems",
            material_type="resume",
        )
        professor_response = self.client.post(
            "/api/professors",
            json={
                "name": "回收站导师",
                "email": "trash-match@example.edu",
                "title": "Professor",
                "university": "Example University",
                "school": "School of Computing",
                "department": "Computer Science",
                "research_direction": "AI agents",
                "recent_papers": ["Agent paper"],
                "profile_url": None,
                "source_url": None,
            },
        )
        self.assertEqual(professor_response.status_code, 201, msg=professor_response.text)
        created = self.client.post(
            "/api/match-analysis-jobs",
            json={
                "identity_id": identity_id,
                "llm_profile_id": llm_id,
                "professor_ids": [professor_response.json()["id"]],
            },
        )
        self.assertEqual(created.status_code, 201, msg=created.text)
        job_id = created.json()["id"]

        blocked = self.client.post(f"/api/match-analysis-jobs/{job_id}/delete")
        self.assertEqual(blocked.status_code, 400)
        self.assertIn("请先中止/取消任务后再删除", blocked.json()["detail"])

        canceled = self.client.post(f"/api/match-analysis-jobs/{job_id}/cancel")
        self.assertEqual(canceled.status_code, 200, msg=canceled.text)
        deleted = self.client.post(f"/api/match-analysis-jobs/{job_id}/delete")
        self.assertEqual(deleted.status_code, 200, msg=deleted.text)
        self.assertIsNotNone(deleted.json()["job"]["deleted_at"])

        repeated_delete = self.client.post(f"/api/match-analysis-jobs/{job_id}/delete")
        self.assertEqual(repeated_delete.status_code, 200, msg=repeated_delete.text)

        current = self.client.get(
            "/api/match-analysis-jobs",
            params={"identity_id": identity_id, "llm_profile_id": llm_id},
        )
        self.assertEqual(current.status_code, 200)
        self.assertEqual(current.json(), [])

        trash = self.client.get(
            "/api/match-analysis-jobs",
            params={"identity_id": identity_id, "llm_profile_id": llm_id, "view": "trash"},
        )
        self.assertEqual(trash.status_code, 200)
        self.assertEqual([item["id"] for item in trash.json()], [job_id])

        restored = self.client.post(f"/api/match-analysis-jobs/{job_id}/restore")
        self.assertEqual(restored.status_code, 200, msg=restored.text)
        self.assertIsNone(restored.json()["job"]["deleted_at"])

        repeated_restore = self.client.post(f"/api/match-analysis-jobs/{job_id}/restore")
        self.assertEqual(repeated_restore.status_code, 200, msg=repeated_restore.text)

    def test_cancel_match_analysis_job(self) -> None:
        identity_id = self._create_identity(with_imap=False)
        llm_id = self._create_llm()
        self._upload_material(
            identity_id,
            filename="resume.txt",
            content=b"AI systems",
            material_type="resume",
        )
        professor_response = self.client.post(
            "/api/professors",
            json={
                "name": "取消任务导师",
                "email": "cancel-job@example.edu",
                "title": "Professor",
                "university": "Example University",
                "school": "School of Computing",
                "department": "Computer Science",
                "research_direction": "AI agents",
                "recent_papers": [],
                "profile_url": None,
                "source_url": None,
            },
        )
        self.assertEqual(professor_response.status_code, 201, msg=professor_response.text)
        professor_id = professor_response.json()["id"]
        created = self.client.post(
            "/api/match-analysis-jobs",
            json={
                "identity_id": identity_id,
                "llm_profile_id": llm_id,
                "professor_ids": [professor_id],
            },
        )
        self.assertEqual(created.status_code, 201, msg=created.text)
        job_id = created.json()["id"]

        canceled = self.client.post(f"/api/match-analysis-jobs/{job_id}/cancel")
        self.assertEqual(canceled.status_code, 200, msg=canceled.text)
        self.assertTrue(canceled.json()["ok"])
        self.assertEqual(canceled.json()["job"]["status"], "canceled")

    def test_retry_failed_match_analysis_job_returns_400_when_no_failed_items(self) -> None:
        identity_id = self._create_identity(with_imap=False)
        llm_id = self._create_llm()
        self._upload_material(
            identity_id,
            filename="resume.txt",
            content=b"AI systems",
            material_type="resume",
        )
        professor_response = self.client.post(
            "/api/professors",
            json={
                "name": "重试任务导师",
                "email": "retry-job@example.edu",
                "title": "Professor",
                "university": "Example University",
                "school": "School of Computing",
                "department": "Computer Science",
                "research_direction": "AI agents",
                "recent_papers": [],
                "profile_url": None,
                "source_url": None,
            },
        )
        self.assertEqual(professor_response.status_code, 201, msg=professor_response.text)
        professor_id = professor_response.json()["id"]
        created = self.client.post(
            "/api/match-analysis-jobs",
            json={
                "identity_id": identity_id,
                "llm_profile_id": llm_id,
                "professor_ids": [professor_id],
            },
        )
        self.assertEqual(created.status_code, 201, msg=created.text)
        job_id = created.json()["id"]

        retried = self.client.post(f"/api/match-analysis-jobs/{job_id}/retry-failed")
        self.assertEqual(retried.status_code, 400)
        self.assertIn("没有可重试的失败项", retried.json()["detail"])

    def test_create_scheduled_batch_task_returns_normalized_scheduled_dates(self) -> None:
        identity_id = self._create_identity(with_imap=False)
        llm_profile_id = self._create_llm()
        self.client.post("/api/professors/import-sample")
        professor_id = self.client.get("/api/professors").json()[0]["id"]

        response = self.client.post(
            "/api/batch-tasks",
            json={
                "identity_id": identity_id,
                "llm_profile_id": llm_profile_id,
                "name": "日历定时发送",
                "professor_ids": [professor_id],
                "schedule_type": "scheduled",
                "scheduled_dates": ["2026-05-04", "2026-04-28", "2026-05-04"],
                "window_start_time": "09:00",
                "window_end_time": "18:00",
                "emails_per_window": 20,
                "primary_material_id": None,
                "email_subject": "Hello {{导师姓名}}",
                "email_body": "Body",
                "selected_material_ids": None,
                "outreach_generation_mode": "template",
                "outreach_template_subject": "Hello {{导师姓名}}",
                "outreach_template_body_text": "Body",
                "outreach_template_body_html": None,
            },
        )

        self.assertEqual(response.status_code, 201)
        self.assertEqual(response.json()["scheduled_dates"], ["2026-04-28", "2026-05-04"])

    def test_create_scheduled_batch_task_rejects_invalid_window_time_format(self) -> None:
        identity_id = self._create_identity(with_imap=False)
        llm_profile_id = self._create_llm()
        self.client.post("/api/professors/import-sample")
        professor_id = self.client.get("/api/professors").json()[0]["id"]

        response = self.client.post(
            "/api/batch-tasks",
            json={
                "identity_id": identity_id,
                "llm_profile_id": llm_profile_id,
                "name": "时间格式校验",
                "professor_ids": [professor_id],
                "schedule_type": "scheduled",
                "scheduled_dates": ["2026-05-04"],
                "window_start_time": "9:00",
                "window_end_time": "18:00",
                "emails_per_window": 20,
                "primary_material_id": None,
                "email_subject": "Hello {{导师姓名}}",
                "email_body": "Body",
                "selected_material_ids": None,
                "outreach_generation_mode": "template",
                "outreach_template_subject": "Hello {{导师姓名}}",
                "outreach_template_body_text": "Body",
                "outreach_template_body_html": None,
            },
        )

        self.assertEqual(response.status_code, 400)
        self.assertIn("HH:mm", response.json()["detail"])

    def test_batch_task_worker_and_workspace_flow(self) -> None:
        identity_id = self._create_identity(with_imap=False)
        llm_id = self._create_llm()
        resume_material_id = self._upload_material(
            identity_id,
            filename="resume.txt",
            content=b"My background covers large language models and IE.",
            material_type="resume",
        )
        publication_material_id = self._upload_material(
            identity_id,
            filename="paper.md",
            content=b"# Publication\nA paper about extraction and LLMs.",
            material_type="publication",
        )

        import_response = self.client.post("/api/professors/import-sample")
        self.assertEqual(import_response.status_code, 200)

        professor_list = self.client.get(
            "/api/professors",
            params={"identity_id": identity_id, "llm_profile_id": llm_id},
        )
        professors = professor_list.json()
        selected_professor_ids = [item["id"] for item in professors[:2]]

        task_response = self.client.post(
            "/api/batch-tasks",
            json={
                "identity_id": identity_id,
                "llm_profile_id": llm_id,
                "name": "首轮联系任务",
                "professor_ids": selected_professor_ids,
                "schedule_type": "immediate",
                "window_start_time": None,
                "window_end_time": None,
                "emails_per_window": None,
                "primary_material_id": resume_material_id,
                "email_subject": "科研交流申请",
                "email_body": "老师您好，这是自定义模板。",
                "selected_material_ids": [publication_material_id],
            },
        )
        self.assertEqual(task_response.status_code, 201)
        self.assertEqual(task_response.json()["pending_generation_count"], 2)

        updated_professors = self.client.get(
            "/api/professors",
            params={"identity_id": identity_id, "llm_profile_id": llm_id},
        ).json()
        selected_professor = next(
            item for item in updated_professors if item["id"] == selected_professor_ids[0]
        )
        self.assertEqual(selected_professor["status"], "preparing")
        self.assertIsNone(selected_professor["match_score"])

        workspace_before = self.client.get(
            f"/api/workspaces/{selected_professor_ids[0]}",
            params={"identity_id": identity_id, "llm_profile_id": llm_id},
        )
        self.assertEqual(workspace_before.status_code, 200)
        task_id = workspace_before.json()["current_task"]["id"]
        self.assertEqual(
            workspace_before.json()["current_task"]["primary_material_id"],
            resume_material_id,
        )
        self.assertEqual(
            workspace_before.json()["current_task"]["selected_material_ids"],
            [publication_material_id],
        )

        with patch(
            "app.services.task_runtime.llm_runtime.generate_match_evaluation",
            AsyncMock(
                return_value=self._build_match_evaluation_result(
                    match_score=93,
                ),
            ),
        ):
            match_workspace = self.client.post(
                f"/api/email-tasks/{task_id}/calculate-match",
            )

        with patch(
            "app.services.task_runtime.llm_runtime.generate_draft_content",
            AsyncMock(
                return_value=self._build_draft_generation_result(
                    subject="更新后的套磁申请",
                    body_text="老师您好，这是切换默认材料后的草稿。",
                    body_html="<p>老师您好，这是切换默认材料后的草稿。</p>",
                    prompt_tokens=612,
                    completion_tokens=248,
                ),
            ),
        ):
            generated_workspace = self.client.post(
                f"/api/email-tasks/{task_id}/generate-draft",
            )

            switched_workspace = self.client.post(
                f"/api/email-tasks/{task_id}/primary-material",
                json={"primary_material_id": publication_material_id},
            )

        self.assertEqual(match_workspace.status_code, 200)
        self.assertEqual(match_workspace.json()["thread"]["current_task"]["status"], "matched")
        self.assertEqual(match_workspace.json()["thread"]["current_task"]["match_score"], 93)
        self.assertEqual(generated_workspace.status_code, 200)
        self.assertEqual(generated_workspace.json()["current_task"]["status"], "review_required")
        self.assertEqual(generated_workspace.json()["current_task"]["generated_subject"], "更新后的套磁申请")
        self.assertEqual(generated_workspace.json()["current_task"]["last_draft_prompt_tokens"], 612)
        self.assertEqual(generated_workspace.json()["current_task"]["last_draft_completion_tokens"], 248)
        self.assertEqual(generated_workspace.json()["current_task"]["last_draft_total_tokens"], 860)
        self.assertGreater(generated_workspace.json()["current_task"]["estimated_prompt_tokens"], 0)
        self.assertEqual(generated_workspace.json()["messages"][-1]["prompt_tokens"], 612)
        self.assertEqual(generated_workspace.json()["messages"][-1]["completion_tokens"], 248)
        self.assertEqual(generated_workspace.json()["messages"][-1]["total_tokens"], 860)
        operation_logs = self.client.get(
            "/api/diagnostics/operation-logs",
            params={"event_name": "email_task.draft_generated"},
        )
        self.assertEqual(operation_logs.status_code, 200, msg=operation_logs.text)
        draft_generated_metadata = operation_logs.json()["items"][0]["metadata"]
        self.assertEqual(draft_generated_metadata["prompt_tokens"], 612)
        self.assertEqual(draft_generated_metadata["completion_tokens"], 248)
        self.assertEqual(draft_generated_metadata["total_tokens"], 860)
        self.assertEqual(switched_workspace.status_code, 200)
        self.assertEqual(switched_workspace.json()["current_task"]["primary_material_id"], publication_material_id)
        self.assertEqual(switched_workspace.json()["current_task"]["status"], "review_required")

        with patch(
            "app.services.task_runtime.mail_runtime.send_email",
            AsyncMock(
                return_value=self._build_send_result(
                    message_id="<manual-send@example.com>",
                    provider_payload={"smtp_host": "smtp.example.com"},
                ),
            ),
        ) as mocked_send:
            workspace_after = self.client.post(
                f"/api/email-tasks/{task_id}/approve-and-send",
                json={
                    "subject": "科研交流申请",
                    "body_text": "老师您好，我希望进一步交流。",
                    "body_html": None,
                    "selected_material_ids": [],
                },
            )
        payload = workspace_after.json()
        self.assertEqual(workspace_after.status_code, 200)
        self.assertEqual(payload["current_task"]["status"], "sent")
        self.assertNotIn("delivery_mode", payload["current_task"])
        self.assertEqual(payload["current_task"]["selected_material_ids"], [])
        self.assertGreaterEqual(len(payload["messages"]), 2)
        self.assertEqual(payload["messages"][-1]["direction"], "sent")
        mocked_send.assert_awaited_once()

    def test_generate_draft_requires_professor_research_direction(self) -> None:
        identity_id = self._create_identity(with_imap=False)
        llm_id = self._create_llm()
        self._upload_material(
            identity_id,
            filename="resume.txt",
            content=b"My background covers agent systems.",
            material_type="resume",
        )

        professor_response = self.client.post(
            "/api/professors",
            json={
                "name": "李老师",
                "email": "li-missing-research@example.edu",
                "title": "Professor",
                "university": "Example University",
                "school": "School of Computing",
                "department": "Computer Science",
                "research_direction": None,
                "recent_papers": ["Agent paper"],
                "profile_url": None,
                "source_url": None,
            },
        )
        self.assertEqual(professor_response.status_code, 201, msg=professor_response.text)
        professor_id = professor_response.json()["id"]

        workspace = self.client.post(
            f"/api/workspaces/{professor_id}/ensure-task",
            params={"identity_id": identity_id, "llm_profile_id": llm_id},
        )
        self.assertEqual(workspace.status_code, 200, msg=workspace.text)
        task_id = workspace.json()["current_task"]["id"]

        with patch(
            "app.services.task_runtime.llm_runtime.generate_draft_content",
            AsyncMock(
                return_value=self._build_draft_generation_result(
                    subject="不应生成的草稿",
                    body_text="这封草稿不应在缺少研究方向时生成。",
                    body_html="<p>这封草稿不应在缺少研究方向时生成。</p>",
                ),
            ),
        ) as mocked_generate:
            response = self.client.post(f"/api/email-tasks/{task_id}/generate-draft")

        self.assertEqual(response.status_code, 400)
        self.assertIn("请先补充导师研究方向", response.json()["detail"])
        mocked_generate.assert_not_awaited()

        refreshed = self.client.get(
            f"/api/workspaces/{professor_id}",
            params={"identity_id": identity_id, "llm_profile_id": llm_id},
        )
        self.assertEqual(refreshed.status_code, 200, msg=refreshed.text)
        self.assertFalse(
            any(message["direction"] == "draft" for message in refreshed.json()["messages"]),
        )

    def test_template_draft_does_not_require_professor_research_direction(self) -> None:
        identity_response = self.client.post(
            "/api/identities",
            json=self._build_identity_payload(
                with_imap=False,
                outreach_generation_mode="template",
                outreach_template_subject="申请与{{name}}老师交流",
                outreach_template_body_text="{{name}}老师您好，我是{{sender_name}}。",
            ),
        )
        self.assertEqual(identity_response.status_code, 201, msg=identity_response.text)
        identity_id = identity_response.json()["id"]
        llm_id = self._create_llm()

        professor_response = self.client.post(
            "/api/professors",
            json={
                "name": "模板模式导师",
                "email": "template-no-research@example.edu",
                "title": "Professor",
                "university": "Example University",
                "school": "School of Computing",
                "department": "Computer Science",
                "research_direction": None,
                "recent_papers": [],
                "profile_url": None,
                "source_url": None,
            },
        )
        self.assertEqual(professor_response.status_code, 201, msg=professor_response.text)
        professor_id = professor_response.json()["id"]

        workspace = self.client.post(
            f"/api/workspaces/{professor_id}/ensure-task",
            params={"identity_id": identity_id, "llm_profile_id": llm_id},
        )
        self.assertEqual(workspace.status_code, 200, msg=workspace.text)
        task_id = workspace.json()["current_task"]["id"]

        with patch(
            "app.services.task_runtime.llm_runtime.generate_draft_content",
            AsyncMock(side_effect=AssertionError("模板模式不应调用 LLM 草稿生成")),
        ) as mocked_generate:
            response = self.client.post(f"/api/email-tasks/{task_id}/generate-draft")

        self.assertEqual(response.status_code, 200, msg=response.text)
        self.assertEqual(response.json()["current_task"]["status"], "review_required")
        self.assertEqual(response.json()["messages"][-1]["direction"], "draft")
        mocked_generate.assert_not_awaited()

    def test_draft_preview_returns_content_without_persisting_changes(self) -> None:
        identity_id = self._create_identity(with_imap=False)
        llm_id = self._create_llm()
        self._upload_material(
            identity_id,
            filename="resume.txt",
            content=b"My background covers agent systems.",
            material_type="resume",
        )

        professor_response = self.client.post(
            "/api/professors",
            json={
                "name": "预览导师",
                "email": "preview-research@example.edu",
                "title": "Professor",
                "university": "Example University",
                "school": "School of Computing",
                "department": "Computer Science",
                "research_direction": "Agent systems",
                "recent_papers": ["Agent paper"],
                "profile_url": None,
                "source_url": None,
            },
        )
        self.assertEqual(professor_response.status_code, 201, msg=professor_response.text)
        professor_id = professor_response.json()["id"]

        workspace = self.client.post(
            f"/api/workspaces/{professor_id}/ensure-task",
            params={"identity_id": identity_id, "llm_profile_id": llm_id},
        )
        self.assertEqual(workspace.status_code, 200, msg=workspace.text)
        task_id = workspace.json()["current_task"]["id"]

        with patch(
            "app.services.task_runtime.llm_runtime.generate_draft_content",
            AsyncMock(
                return_value=self._build_draft_generation_result(
                    subject="预览主题",
                    body_text="预览正文",
                    body_html="<p>预览正文</p>",
                    prompt_tokens=120,
                    completion_tokens=30,
                ),
            ),
        ) as mocked_generate:
            response = self.client.post(f"/api/email-tasks/{task_id}/draft-preview")

        self.assertEqual(response.status_code, 200, msg=response.text)
        preview = response.json()
        self.assertEqual(preview["subject"], "预览主题")
        self.assertEqual(preview["body_text"], "预览正文")
        self.assertEqual(preview["usage"]["prompt_tokens"], 120)
        mocked_generate.assert_awaited_once()

        refreshed = self.client.get(
            f"/api/workspaces/{professor_id}",
            params={"identity_id": identity_id, "llm_profile_id": llm_id},
        )
        self.assertEqual(refreshed.status_code, 200, msg=refreshed.text)
        self.assertIsNone(refreshed.json()["current_task"]["generated_subject"])
        self.assertFalse(
            any(message["direction"] == "draft" for message in refreshed.json()["messages"]),
        )

    def test_calculate_match_keeps_low_score_task_in_matched_state(self) -> None:
        identity_id = self._create_identity(with_imap=False)
        llm_id = self._create_llm()
        material_id = self._upload_material(
            identity_id,
            filename="resume.txt",
            content=b"My background covers information extraction and agents.",
            material_type="resume",
        )
        set_primary_response = self.client.post(f"/api/materials/{material_id}/set-primary")
        self.assertEqual(set_primary_response.status_code, 200, msg=set_primary_response.text)

        professor_response = self.client.post(
            "/api/professors",
            json={
                "name": "低分匹配导师",
                "email": "low-score-match@example.edu",
                "title": "Professor",
                "university": "Example University",
                "school": "School of Computing",
                "department": "Computer Science",
                "research_direction": "Agents",
                "recent_papers": [],
                "profile_url": None,
                "source_url": None,
            },
        )
        self.assertEqual(professor_response.status_code, 201, msg=professor_response.text)
        professor_id = professor_response.json()["id"]

        connection = sqlite3.connect(self.db_path)
        try:
            connection.execute(
                """
                UPDATE identity_profiles
                SET match_threshold = ?
                WHERE id = ?
                """,
                (95, identity_id),
            )
            connection.commit()
        finally:
            connection.close()

        ensure_response = self.client.post(
            f"/api/workspaces/{professor_id}/ensure-task",
            params={"identity_id": identity_id, "llm_profile_id": llm_id},
        )
        self.assertEqual(ensure_response.status_code, 200, msg=ensure_response.text)
        task_id = ensure_response.json()["current_task"]["id"]

        with patch(
            "app.services.task_runtime.llm_runtime.generate_match_evaluation",
            AsyncMock(return_value=self._build_match_evaluation_result(match_score=18)),
        ):
            response = self.client.post(f"/api/email-tasks/{task_id}/calculate-match")

        self.assertEqual(response.status_code, 200, msg=response.text)
        self.assertEqual(response.json()["thread"]["current_task"]["match_score"], 18)
        self.assertEqual(response.json()["thread"]["current_task"]["status"], "matched")

    def test_calculate_match_requires_professor_research_evidence(self) -> None:
        identity_id = self._create_identity(with_imap=False)
        llm_id = self._create_llm()
        material_id = self._upload_material(
            identity_id,
            filename="resume.txt",
            content=b"My background covers information extraction and agents.",
            material_type="resume",
        )
        set_primary_response = self.client.post(f"/api/materials/{material_id}/set-primary")
        self.assertEqual(set_primary_response.status_code, 200, msg=set_primary_response.text)

        professor_response = self.client.post(
            "/api/professors",
            json={
                "name": "缺少研究信息导师",
                "email": "missing-evidence@example.edu",
                "title": "Professor",
                "university": "Example University",
                "school": "School of AI",
                "department": "Computer Science",
                "research_direction": None,
                "recent_papers": [],
                "profile_url": None,
                "source_url": None,
            },
        )
        self.assertEqual(professor_response.status_code, 201, msg=professor_response.text)
        professor_id = professor_response.json()["id"]

        ensure_response = self.client.post(
            f"/api/workspaces/{professor_id}/ensure-task",
            params={"identity_id": identity_id, "llm_profile_id": llm_id},
        )
        self.assertEqual(ensure_response.status_code, 200, msg=ensure_response.text)
        task_id = ensure_response.json()["current_task"]["id"]

        with patch(
            "app.services.task_runtime.llm_runtime.generate_match_evaluation",
            AsyncMock(side_effect=AssertionError("不应在缺少研究信息时调用模型")),
        ):
            response = self.client.post(f"/api/email-tasks/{task_id}/calculate-match")

        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.json()["detail"], "缺少研究方向或近期论文，暂不能分析匹配度")

    def test_calculate_match_returns_409_when_run_is_already_running(self) -> None:
        from app.services.task_runtime import MatchAnalysisAlreadyRunningError

        with patch(
            "app.api.email_tasks.calculate_task_match_once",
            AsyncMock(side_effect=MatchAnalysisAlreadyRunningError("该任务正在分析中")),
        ):
            response = self.client.post("/api/email-tasks/1/calculate-match")

        self.assertEqual(response.status_code, 409)
        self.assertEqual(response.json()["detail"], "该任务正在分析中")

    def test_workspace_thread_includes_professor_recent_papers(self) -> None:
        identity_id = self._create_identity(with_imap=False)
        llm_id = self._create_llm()

        professor_response = self.client.post(
            "/api/professors",
            json={
                "name": "只有论文导师",
                "email": "paper-only@example.edu",
                "title": "Professor",
                "university": "Example University",
                "school": "School of AI",
                "department": "Computer Science",
                "research_direction": None,
                "recent_papers": ["Paper Evidence"],
                "profile_url": None,
                "source_url": None,
            },
        )
        self.assertEqual(professor_response.status_code, 201, msg=professor_response.text)
        professor_id = professor_response.json()["id"]

        response = self.client.post(
            f"/api/workspaces/{professor_id}/ensure-task",
            params={"identity_id": identity_id, "llm_profile_id": llm_id},
        )

        self.assertEqual(response.status_code, 200, msg=response.text)
        self.assertEqual(response.json()["professor"]["recent_papers"], ["Paper Evidence"])

    def test_stop_batch_task_marks_pending_items_as_canceled(self) -> None:
        identity_id = self._create_identity(with_imap=False)
        llm_id = self._create_llm()

        self.client.post("/api/professors/import-sample")
        professor_ids = [item["id"] for item in self.client.get("/api/professors").json()[:3]]

        create_response = self.client.post(
            "/api/batch-tasks",
            json={
                "identity_id": identity_id,
                "llm_profile_id": llm_id,
                "name": "停止后取消未完成任务",
                "professor_ids": professor_ids,
                "schedule_type": "immediate",
                "window_start_time": None,
                "window_end_time": None,
                "emails_per_window": None,
                "primary_material_id": None,
                "email_subject": "联系 {{name}}",
                "email_body": "老师您好，我是{{sender_name}}。",
                "selected_material_ids": None,
            },
        )
        self.assertEqual(create_response.status_code, 201, msg=create_response.text)
        batch_task_id = create_response.json()["id"]

        connection = sqlite3.connect(self.db_path)
        try:
            task_ids = [
                row[0]
                for row in connection.execute(
                    """
                    SELECT id
                    FROM email_tasks
                    WHERE batch_task_id = ?
                    ORDER BY id
                    """,
                    (batch_task_id,),
                ).fetchall()
            ]
            self.assertEqual(len(task_ids), 3)
            connection.execute(
                "UPDATE email_tasks SET status = ? WHERE id = ?",
                ("matched", task_ids[0]),
            )
            connection.execute(
                "UPDATE email_tasks SET status = ? WHERE id = ?",
                ("review_required", task_ids[1]),
            )
            connection.execute(
                """
                UPDATE email_tasks
                SET status = ?, sent_at = CURRENT_TIMESTAMP
                WHERE id = ?
                """,
                ("sent", task_ids[2]),
            )
            connection.commit()
        finally:
            connection.close()

        stop_response = self.client.post(f"/api/batch-tasks/{batch_task_id}/stop")
        self.assertEqual(stop_response.status_code, 200, msg=stop_response.text)

        connection = sqlite3.connect(self.db_path)
        try:
            rows = connection.execute(
                """
                SELECT id, status, cancellation_reason
                FROM email_tasks
                WHERE batch_task_id = ?
                ORDER BY id
                """,
                (batch_task_id,),
            ).fetchall()
        finally:
            connection.close()

        self.assertEqual(
            rows,
            [
                (task_ids[0], "canceled", "batch_stopped"),
                (task_ids[1], "canceled", "batch_stopped"),
                (task_ids[2], "sent", None),
            ],
        )

    def test_stop_batch_task_keeps_send_failed_items_failed(self) -> None:
        identity_id = self._create_identity(with_imap=False)
        llm_id = self._create_llm()

        self.client.post("/api/professors/import-sample")
        professor_ids = [item["id"] for item in self.client.get("/api/professors").json()[:2]]

        create_response = self.client.post(
            "/api/batch-tasks",
            json={
                "identity_id": identity_id,
                "llm_profile_id": llm_id,
                "name": "停止时保留失败任务",
                "professor_ids": professor_ids,
                "schedule_type": "immediate",
                "window_start_time": None,
                "window_end_time": None,
                "emails_per_window": None,
                "primary_material_id": None,
                "email_subject": "联系 {{name}}",
                "email_body": "老师您好，我是{{sender_name}}。",
                "selected_material_ids": None,
            },
        )
        self.assertEqual(create_response.status_code, 201, msg=create_response.text)
        batch_task_id = create_response.json()["id"]

        connection = sqlite3.connect(self.db_path)
        try:
            task_ids = [
                row[0]
                for row in connection.execute(
                    """
                    SELECT id
                    FROM email_tasks
                    WHERE batch_task_id = ?
                    ORDER BY id
                    """,
                    (batch_task_id,),
                ).fetchall()
            ]
            self.assertEqual(len(task_ids), 2)
            connection.execute(
                """
                UPDATE email_tasks
                SET status = ?, last_error = ?
                WHERE id = ?
                """,
                ("send_failed", "smtp timeout", task_ids[0]),
            )
            connection.execute(
                "UPDATE email_tasks SET status = ? WHERE id = ?",
                ("matched", task_ids[1]),
            )
            connection.commit()
        finally:
            connection.close()

        stop_response = self.client.post(f"/api/batch-tasks/{batch_task_id}/stop")
        self.assertEqual(stop_response.status_code, 200, msg=stop_response.text)

        connection = sqlite3.connect(self.db_path)
        try:
            rows = connection.execute(
                """
                SELECT id, status, cancellation_reason, last_error
                FROM email_tasks
                WHERE batch_task_id = ?
                ORDER BY id
                """,
                (batch_task_id,),
            ).fetchall()
        finally:
            connection.close()

        self.assertEqual(
            rows,
            [
                (task_ids[0], "send_failed", None, "smtp timeout"),
                (task_ids[1], "canceled", "batch_stopped", None),
            ],
        )

    def test_continue_manually_creates_manual_child_task_from_batch_stopped_task(self) -> None:
        identity_id = self._create_identity(with_imap=False)
        llm_id = self._create_llm()
        primary_material_id = self._upload_material(
            identity_id,
            filename="resume.txt",
            content=b"My background covers agent systems and information extraction.",
            material_type="resume",
        )
        attachment_material_id = self._upload_material(
            identity_id,
            filename="paper.pdf",
            content=b"%PDF-1.4 test attachment",
            material_type="portfolio",
        )
        set_primary_response = self.client.post(f"/api/materials/{primary_material_id}/set-primary")
        self.assertEqual(set_primary_response.status_code, 200, msg=set_primary_response.text)

        professor_response = self.client.post(
            "/api/professors",
            json={
                "name": "手动继续联系导师",
                "email": "continue-manually@example.edu",
                "title": "Professor",
                "university": "Example University",
                "school": "School of Computing",
                "department": "Computer Science",
                "research_direction": "Agent systems",
                "recent_papers": [],
                "profile_url": None,
                "source_url": None,
            },
        )
        self.assertEqual(professor_response.status_code, 201, msg=professor_response.text)
        professor_id = professor_response.json()["id"]

        create_response = self.client.post(
            "/api/batch-tasks",
            json={
                "identity_id": identity_id,
                "llm_profile_id": llm_id,
                "name": "继续联系批量任务",
                "professor_ids": [professor_id],
                "schedule_type": "immediate",
                "window_start_time": None,
                "window_end_time": None,
                "emails_per_window": None,
                "primary_material_id": primary_material_id,
                "email_subject": None,
                "email_body": None,
                "selected_material_ids": [attachment_material_id],
                "outreach_generation_mode": "template",
                "outreach_template_subject": "继续联系 {{name}}",
                "outreach_template_body_text": "继续联系正文 {{name}}",
                "outreach_template_body_html": "<p>继续联系正文 {{name}}</p>",
            },
        )
        self.assertEqual(create_response.status_code, 201, msg=create_response.text)
        batch_task_id = create_response.json()["id"]

        connection = sqlite3.connect(self.db_path)
        try:
            parent_task_id = connection.execute(
                """
                SELECT id
                FROM email_tasks
                WHERE batch_task_id = ?
                """,
                (batch_task_id,),
            ).fetchone()[0]
            connection.execute(
                """
                UPDATE email_tasks
                SET status = ?,
                    cancellation_reason = ?,
                    match_score = ?,
                    match_reason = ?,
                    fit_points = ?,
                    risk_points = ?,
                    match_keywords = ?,
                    generated_subject = ?,
                    generated_content_text = ?,
                    generated_content_html = ?,
                    selected_material_ids = ?,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
                """,
                (
                    "canceled",
                    "batch_stopped",
                    91,
                    "研究方向与材料高度匹配",
                    json.dumps(["研究方向契合"]),
                    json.dumps(["需要补充近期成果"]),
                    json.dumps(["agent"]),
                    "旧草稿主题",
                    "旧草稿正文",
                    "<p>旧草稿正文</p>",
                    json.dumps([attachment_material_id]),
                    parent_task_id,
                ),
            )
            connection.commit()
        finally:
            connection.close()

        workspace_before = self.client.get(
            f"/api/workspaces/{professor_id}",
            params={"identity_id": identity_id, "llm_profile_id": llm_id},
        )
        self.assertEqual(workspace_before.status_code, 200, msg=workspace_before.text)
        before_task = workspace_before.json()["current_task"]
        self.assertEqual(before_task["id"], parent_task_id)
        self.assertEqual(before_task["source"], "batch")
        self.assertIsNone(before_task["parent_task_id"])
        self.assertEqual(before_task["cancellation_reason"], "batch_stopped")
        self.assertTrue(before_task["can_continue_manually"])
        self.assertFalse(before_task["can_write_follow_up"])

        response = self.client.post(f"/api/email-tasks/{parent_task_id}/continue-manually")

        self.assertEqual(response.status_code, 200, msg=response.text)
        payload = response.json()
        current_task = payload["current_task"]
        self.assertNotEqual(current_task["id"], parent_task_id)
        self.assertIsNone(current_task["batch_task_id"])
        self.assertEqual(current_task["source"], "manual")
        self.assertEqual(current_task["parent_task_id"], parent_task_id)
        self.assertEqual(current_task["status"], "review_required")
        self.assertIsNone(current_task["cancellation_reason"])
        self.assertEqual(current_task["primary_material_id"], primary_material_id)
        self.assertEqual(current_task["selected_material_ids"], [attachment_material_id])
        self.assertEqual(current_task["match_score"], 91)
        self.assertEqual(current_task["match_reason"], "研究方向与材料高度匹配")
        self.assertEqual(current_task["fit_points"], ["研究方向契合"])
        self.assertEqual(current_task["risk_points"], ["需要补充近期成果"])
        self.assertEqual(current_task["match_keywords"], ["agent"])
        self.assertEqual(current_task["generated_subject"], "旧草稿主题")
        self.assertEqual(current_task["generated_content_text"], "旧草稿正文")
        self.assertEqual(current_task["generated_content_html"], "<p>旧草稿正文</p>")
        self.assertEqual(current_task["outreach_generation_mode"], "template")
        self.assertEqual(current_task["outreach_template_subject"], "继续联系 {{name}}")
        self.assertEqual(current_task["outreach_template_body_text"], "继续联系正文 {{name}}")
        self.assertEqual(current_task["outreach_template_body_html"], "<p>继续联系正文 {{name}}</p>")
        self.assertFalse(current_task["can_continue_manually"])
        self.assertFalse(current_task["can_write_follow_up"])

        connection = sqlite3.connect(self.db_path)
        try:
            rows = connection.execute(
                """
                SELECT id, source, batch_task_id, parent_task_id, status, cancellation_reason,
                       generated_subject, generated_content_text, selected_material_ids
                FROM email_tasks
                WHERE professor_id = ?
                ORDER BY id
                """,
                (professor_id,),
            ).fetchall()
        finally:
            connection.close()

        self.assertEqual(
            rows,
            [
                (
                    parent_task_id,
                    "batch",
                    batch_task_id,
                    None,
                    "canceled",
                    "batch_stopped",
                    "旧草稿主题",
                    "旧草稿正文",
                    json.dumps([attachment_material_id]),
                ),
                (
                    current_task["id"],
                    "manual",
                    None,
                    parent_task_id,
                    "review_required",
                    None,
                    "旧草稿主题",
                    "旧草稿正文",
                    json.dumps([attachment_material_id]),
                ),
            ],
        )

    def test_continue_manually_restores_matched_when_parent_has_match_without_draft(self) -> None:
        identity_id = self._create_identity(with_imap=False)
        llm_id = self._create_llm()
        primary_material_id = self._upload_material(
            identity_id,
            filename="resume.txt",
            content=b"My background covers agent systems and information extraction.",
            material_type="resume",
        )
        set_primary_response = self.client.post(f"/api/materials/{primary_material_id}/set-primary")
        self.assertEqual(set_primary_response.status_code, 200, msg=set_primary_response.text)

        professor_response = self.client.post(
            "/api/professors",
            json={
                "name": "无草稿已匹配导师",
                "email": "continue-matched@example.edu",
                "title": "Professor",
                "university": "Example University",
                "school": "School of Computing",
                "department": "Computer Science",
                "research_direction": "Agent systems",
                "recent_papers": [],
                "profile_url": None,
                "source_url": None,
            },
        )
        self.assertEqual(professor_response.status_code, 201, msg=professor_response.text)
        professor_id = professor_response.json()["id"]

        create_response = self.client.post(
            "/api/batch-tasks",
            json={
                "identity_id": identity_id,
                "llm_profile_id": llm_id,
                "name": "继续联系匹配分支",
                "professor_ids": [professor_id],
                "schedule_type": "immediate",
                "window_start_time": None,
                "window_end_time": None,
                "emails_per_window": None,
                "primary_material_id": primary_material_id,
                "email_subject": "联系 {{name}}",
                "email_body": "联系正文 {{name}}",
                "selected_material_ids": None,
            },
        )
        self.assertEqual(create_response.status_code, 201, msg=create_response.text)
        batch_task_id = create_response.json()["id"]

        connection = sqlite3.connect(self.db_path)
        try:
            parent_task_id = connection.execute(
                "SELECT id FROM email_tasks WHERE batch_task_id = ?",
                (batch_task_id,),
            ).fetchone()[0]
            connection.execute(
                """
                UPDATE email_tasks
                SET status = ?,
                    cancellation_reason = ?,
                    match_score = ?,
                    match_reason = ?,
                    fit_points = ?,
                    risk_points = ?,
                    match_keywords = ?,
                    generated_subject = NULL,
                    generated_content_text = NULL,
                    generated_content_html = NULL,
                    approved_subject = NULL,
                    approved_body_text = NULL,
                    approved_body_html = NULL
                WHERE id = ?
                """,
                (
                    "canceled",
                    "batch_stopped",
                    75,
                    "匹配结果仍可复用",
                    json.dumps(["方向契合"]),
                    json.dumps(["需补充研究计划"]),
                    json.dumps(["match"]),
                    parent_task_id,
                ),
            )
            connection.commit()
        finally:
            connection.close()

        response = self.client.post(f"/api/email-tasks/{parent_task_id}/continue-manually")

        self.assertEqual(response.status_code, 200, msg=response.text)
        current_task = response.json()["current_task"]
        self.assertEqual(current_task["status"], "matched")
        self.assertEqual(current_task["parent_task_id"], parent_task_id)
        self.assertIsNone(current_task["generated_subject"])
        self.assertIsNone(current_task["generated_content_text"])
        self.assertIsNone(current_task["generated_content_html"])
        self.assertEqual(current_task["match_score"], 75)
        self.assertEqual(current_task["match_reason"], "匹配结果仍可复用")

    def test_continue_manually_restores_discovered_when_parent_has_no_match(self) -> None:
        identity_id = self._create_identity(with_imap=False)
        llm_id = self._create_llm()

        professor_response = self.client.post(
            "/api/professors",
            json={
                "name": "未匹配导师",
                "email": "continue-discovered@example.edu",
                "title": "Professor",
                "university": "Example University",
                "school": "School of Computing",
                "department": "Computer Science",
                "research_direction": "Agent systems",
                "recent_papers": [],
                "profile_url": None,
                "source_url": None,
            },
        )
        self.assertEqual(professor_response.status_code, 201, msg=professor_response.text)
        professor_id = professor_response.json()["id"]

        create_response = self.client.post(
            "/api/batch-tasks",
            json={
                "identity_id": identity_id,
                "llm_profile_id": llm_id,
                "name": "继续联系 discovered 分支",
                "professor_ids": [professor_id],
                "schedule_type": "immediate",
                "window_start_time": None,
                "window_end_time": None,
                "emails_per_window": None,
                "primary_material_id": None,
                "email_subject": "联系 {{name}}",
                "email_body": "联系正文 {{name}}",
                "selected_material_ids": None,
            },
        )
        self.assertEqual(create_response.status_code, 201, msg=create_response.text)
        batch_task_id = create_response.json()["id"]

        connection = sqlite3.connect(self.db_path)
        try:
            parent_task_id = connection.execute(
                "SELECT id FROM email_tasks WHERE batch_task_id = ?",
                (batch_task_id,),
            ).fetchone()[0]
            connection.execute(
                """
                UPDATE email_tasks
                SET status = ?,
                    cancellation_reason = ?,
                    match_score = NULL,
                    match_reason = NULL,
                    fit_points = NULL,
                    risk_points = NULL,
                    match_keywords = NULL,
                    generated_subject = NULL,
                    generated_content_text = NULL,
                    generated_content_html = NULL,
                    approved_subject = NULL,
                    approved_body_text = NULL,
                    approved_body_html = NULL
                WHERE id = ?
                """,
                ("canceled", "batch_stopped", parent_task_id),
            )
            connection.commit()
        finally:
            connection.close()

        response = self.client.post(f"/api/email-tasks/{parent_task_id}/continue-manually")

        self.assertEqual(response.status_code, 200, msg=response.text)
        current_task = response.json()["current_task"]
        self.assertEqual(current_task["status"], "discovered")
        self.assertEqual(current_task["parent_task_id"], parent_task_id)
        self.assertIsNone(current_task["match_score"])
        self.assertIsNone(current_task["match_reason"])
        self.assertIsNone(current_task["generated_subject"])

    def test_continue_manually_rejects_duplicate_manual_child_creation(self) -> None:
        identity_id = self._create_identity(with_imap=False)
        llm_id = self._create_llm()

        professor_response = self.client.post(
            "/api/professors",
            json={
                "name": "继续联系重复派生导师",
                "email": "continue-duplicate@example.edu",
                "title": "Professor",
                "university": "Example University",
                "school": "School of Computing",
                "department": "Computer Science",
                "research_direction": "Agent systems",
                "recent_papers": [],
                "profile_url": None,
                "source_url": None,
            },
        )
        self.assertEqual(professor_response.status_code, 201, msg=professor_response.text)
        professor_id = professor_response.json()["id"]

        create_response = self.client.post(
            "/api/batch-tasks",
            json={
                "identity_id": identity_id,
                "llm_profile_id": llm_id,
                "name": "继续联系重复派生",
                "professor_ids": [professor_id],
                "schedule_type": "immediate",
                "window_start_time": None,
                "window_end_time": None,
                "emails_per_window": None,
                "primary_material_id": None,
                "email_subject": "联系 {{name}}",
                "email_body": "联系正文 {{name}}",
                "selected_material_ids": None,
            },
        )
        self.assertEqual(create_response.status_code, 201, msg=create_response.text)
        batch_task_id = create_response.json()["id"]

        connection = sqlite3.connect(self.db_path)
        try:
            parent_task_id = connection.execute(
                "SELECT id FROM email_tasks WHERE batch_task_id = ?",
                (batch_task_id,),
            ).fetchone()[0]
            connection.execute(
                """
                UPDATE email_tasks
                SET status = ?, cancellation_reason = ?
                WHERE id = ?
                """,
                ("canceled", "batch_stopped", parent_task_id),
            )
            connection.commit()
        finally:
            connection.close()

        first_response = self.client.post(f"/api/email-tasks/{parent_task_id}/continue-manually")
        second_response = self.client.post(f"/api/email-tasks/{parent_task_id}/continue-manually")

        self.assertEqual(first_response.status_code, 200, msg=first_response.text)
        self.assertEqual(second_response.status_code, 400, msg=second_response.text)

    def test_continue_manually_returns_404_for_missing_task(self) -> None:
        response = self.client.post("/api/email-tasks/9999/continue-manually")

        self.assertEqual(response.status_code, 404, msg=response.text)
        self.assertEqual(response.json()["detail"], "EmailTask 9999 不存在")

    def test_continue_manually_rejects_task_without_canceled_batch_stopped_guard(self) -> None:
        identity_id = self._create_identity(with_imap=False)
        llm_id = self._create_llm()

        professor_response = self.client.post(
            "/api/professors",
            json={
                "name": "继续联系非法状态导师",
                "email": "continue-guard@example.edu",
                "title": "Professor",
                "university": "Example University",
                "school": "School of Computing",
                "department": "Computer Science",
                "research_direction": "Agent systems",
                "recent_papers": [],
                "profile_url": None,
                "source_url": None,
            },
        )
        self.assertEqual(professor_response.status_code, 201, msg=professor_response.text)
        professor_id = professor_response.json()["id"]

        ensure_response = self.client.post(
            f"/api/workspaces/{professor_id}/ensure-task",
            params={"identity_id": identity_id, "llm_profile_id": llm_id},
        )
        self.assertEqual(ensure_response.status_code, 200, msg=ensure_response.text)
        task_id = ensure_response.json()["current_task"]["id"]

        cases = [
            ("matched", None),
            ("canceled", None),
        ]
        for status_value, cancellation_reason in cases:
            with self.subTest(status=status_value, cancellation_reason=cancellation_reason):
                connection = sqlite3.connect(self.db_path)
                try:
                    connection.execute(
                        """
                        UPDATE email_tasks
                        SET status = ?, cancellation_reason = ?
                        WHERE id = ?
                        """,
                        (status_value, cancellation_reason, task_id),
                    )
                    connection.commit()
                finally:
                    connection.close()

                response = self.client.post(f"/api/email-tasks/{task_id}/continue-manually")

                self.assertEqual(response.status_code, 400, msg=response.text)

    def test_approve_and_send_rejects_canceled_batch_stopped_parent_task(self) -> None:
        task_id = self._create_canceled_batch_stopped_parent_task(
            email="approve-send-guard@example.edu",
        )

        with patch(
            "app.services.task_runtime.mail_runtime.send_email",
            AsyncMock(
                return_value=self._build_send_result(
                    message_id="<guarded-send@example.com>",
                    provider_payload={"smtp_host": "smtp.example.com"},
                ),
            ),
        ) as mocked_send:
            response = self.client.post(
                f"/api/email-tasks/{task_id}/approve-and-send",
                json={
                    "subject": "直接发送",
                    "body_text": "老师您好，这里尝试直接发送。",
                    "body_html": None,
                    "selected_material_ids": [],
                },
            )

        self.assertEqual(response.status_code, 400, msg=response.text)
        self.assertEqual(
            response.json()["detail"],
            "该任务已因批量任务停止而取消，请先“作为单独联系继续”后再执行此操作",
        )
        mocked_send.assert_not_awaited()

    def test_approve_and_schedule_rejects_canceled_batch_stopped_parent_task(self) -> None:
        task_id = self._create_canceled_batch_stopped_parent_task(
            email="approve-schedule-guard@example.edu",
        )

        response = self.client.post(
            f"/api/email-tasks/{task_id}/approve-and-schedule",
            json={
                "subject": "稍后发送",
                "body_text": "老师您好，这里尝试直接定时发送。",
                "body_html": None,
                "selected_material_ids": [],
                "scheduled_at": (datetime.now(UTC) + timedelta(hours=1)).isoformat(),
            },
        )

        self.assertEqual(response.status_code, 400, msg=response.text)
        self.assertEqual(
            response.json()["detail"],
            "该任务已因批量任务停止而取消，请先“作为单独联系继续”后再执行此操作",
        )

    def test_start_follow_up_creates_manual_child_task_from_sent_task(self) -> None:
        identity_id = self._create_identity(with_imap=False)
        llm_id = self._create_llm()
        primary_material_id = self._upload_material(
            identity_id,
            filename="resume.txt",
            content=b"My background covers agent systems and information extraction.",
            material_type="resume",
        )
        attachment_material_id = self._upload_material(
            identity_id,
            filename="portfolio.pdf",
            content=b"%PDF-1.4 follow up attachment",
            material_type="portfolio",
        )
        set_primary_response = self.client.post(f"/api/materials/{primary_material_id}/set-primary")
        self.assertEqual(set_primary_response.status_code, 200, msg=set_primary_response.text)

        professor_response = self.client.post(
            "/api/professors",
            json={
                "name": "跟进邮件导师",
                "email": "follow-up@example.edu",
                "title": "Professor",
                "university": "Example University",
                "school": "School of Computing",
                "department": "Computer Science",
                "research_direction": "Information extraction",
                "recent_papers": [],
                "profile_url": None,
                "source_url": None,
            },
        )
        self.assertEqual(professor_response.status_code, 201, msg=professor_response.text)
        professor_id = professor_response.json()["id"]

        ensure_response = self.client.post(
            f"/api/workspaces/{professor_id}/ensure-task",
            params={"identity_id": identity_id, "llm_profile_id": llm_id},
        )
        self.assertEqual(ensure_response.status_code, 200, msg=ensure_response.text)
        parent_task_id = ensure_response.json()["current_task"]["id"]

        connection = sqlite3.connect(self.db_path)
        try:
            connection.execute(
                """
                UPDATE email_tasks
                SET status = ?,
                    source = ?,
                    primary_material_id = ?,
                    match_score = ?,
                    match_reason = ?,
                    fit_points = ?,
                    risk_points = ?,
                    match_keywords = ?,
                    generated_subject = ?,
                    generated_content_text = ?,
                    generated_content_html = ?,
                    approved_subject = ?,
                    approved_body_text = ?,
                    approved_body_html = ?,
                    outreach_generation_mode = ?,
                    outreach_template_subject = ?,
                    outreach_template_body_text = ?,
                    outreach_template_body_html = ?,
                    selected_material_ids = ?,
                    sent_at = CURRENT_TIMESTAMP,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
                """,
                (
                    "sent",
                    "manual",
                    primary_material_id,
                    88,
                    "已建立初步联系",
                    json.dumps(["研究主题重合"]),
                    json.dumps(["需要补充最新进展"]),
                    json.dumps(["nlp"]),
                    "历史草稿主题",
                    "历史草稿正文",
                    "<p>历史草稿正文</p>",
                    "已发送主题",
                    "已发送正文",
                    "<p>已发送正文</p>",
                    "template",
                    "跟进主题 {{name}}",
                    "跟进正文 {{name}}",
                    "<p>跟进正文 {{name}}</p>",
                    json.dumps([attachment_material_id]),
                    parent_task_id,
                ),
            )
            connection.commit()
        finally:
            connection.close()

        workspace_before = self.client.get(
            f"/api/workspaces/{professor_id}",
            params={"identity_id": identity_id, "llm_profile_id": llm_id},
        )
        self.assertEqual(workspace_before.status_code, 200, msg=workspace_before.text)
        before_task = workspace_before.json()["current_task"]
        self.assertEqual(before_task["id"], parent_task_id)
        self.assertEqual(before_task["source"], "manual")
        self.assertIsNone(before_task["parent_task_id"])
        self.assertIsNone(before_task["cancellation_reason"])
        self.assertFalse(before_task["can_continue_manually"])
        self.assertTrue(before_task["can_write_follow_up"])

        response = self.client.post(f"/api/email-tasks/{parent_task_id}/start-follow-up")

        self.assertEqual(response.status_code, 200, msg=response.text)
        payload = response.json()
        current_task = payload["current_task"]
        self.assertNotEqual(current_task["id"], parent_task_id)
        self.assertIsNone(current_task["batch_task_id"])
        self.assertEqual(current_task["source"], "manual")
        self.assertEqual(current_task["parent_task_id"], parent_task_id)
        self.assertEqual(current_task["status"], "matched")
        self.assertIsNone(current_task["cancellation_reason"])
        self.assertEqual(current_task["primary_material_id"], primary_material_id)
        self.assertEqual(current_task["selected_material_ids"], [attachment_material_id])
        self.assertEqual(current_task["match_score"], 88)
        self.assertEqual(current_task["match_reason"], "已建立初步联系")
        self.assertEqual(current_task["fit_points"], ["研究主题重合"])
        self.assertEqual(current_task["risk_points"], ["需要补充最新进展"])
        self.assertEqual(current_task["match_keywords"], ["nlp"])
        self.assertIsNone(current_task["generated_subject"])
        self.assertIsNone(current_task["generated_content_text"])
        self.assertIsNone(current_task["generated_content_html"])
        self.assertIsNone(current_task["approved_subject"])
        self.assertIsNone(current_task["approved_body_text"])
        self.assertIsNone(current_task["approved_body_html"])
        self.assertEqual(current_task["outreach_generation_mode"], "template")
        self.assertEqual(current_task["outreach_template_subject"], "跟进主题 {{name}}")
        self.assertEqual(current_task["outreach_template_body_text"], "跟进正文 {{name}}")
        self.assertEqual(current_task["outreach_template_body_html"], "<p>跟进正文 {{name}}</p>")
        self.assertFalse(current_task["can_continue_manually"])
        self.assertFalse(current_task["can_write_follow_up"])

        connection = sqlite3.connect(self.db_path)
        try:
            rows = connection.execute(
                """
                SELECT id, source, batch_task_id, parent_task_id, status, cancellation_reason,
                       generated_subject, generated_content_text, approved_subject, approved_body_text,
                       selected_material_ids
                FROM email_tasks
                WHERE professor_id = ?
                ORDER BY id
                """,
                (professor_id,),
            ).fetchall()
        finally:
            connection.close()

        self.assertEqual(
            rows,
            [
                (
                    parent_task_id,
                    "manual",
                    None,
                    None,
                    "sent",
                    None,
                    "历史草稿主题",
                    "历史草稿正文",
                    "已发送主题",
                    "已发送正文",
                    json.dumps([attachment_material_id]),
                ),
                (
                    current_task["id"],
                    "manual",
                    None,
                    parent_task_id,
                    "matched",
                    None,
                    None,
                    None,
                    None,
                    None,
                    json.dumps([attachment_material_id]),
                ),
            ],
        )

    def test_start_follow_up_restores_matched_when_parent_has_no_match_result(self) -> None:
        identity_id = self._create_identity(with_imap=False)
        llm_id = self._create_llm()

        professor_response = self.client.post(
            "/api/professors",
            json={
                "name": "无匹配结果跟进导师",
                "email": "follow-up-minimum-status@example.edu",
                "title": "Professor",
                "university": "Example University",
                "school": "School of Computing",
                "department": "Computer Science",
                "research_direction": "Information extraction",
                "recent_papers": [],
                "profile_url": None,
                "source_url": None,
            },
        )
        self.assertEqual(professor_response.status_code, 201, msg=professor_response.text)
        professor_id = professor_response.json()["id"]

        ensure_response = self.client.post(
            f"/api/workspaces/{professor_id}/ensure-task",
            params={"identity_id": identity_id, "llm_profile_id": llm_id},
        )
        self.assertEqual(ensure_response.status_code, 200, msg=ensure_response.text)
        parent_task_id = ensure_response.json()["current_task"]["id"]

        connection = sqlite3.connect(self.db_path)
        try:
            connection.execute(
                """
                UPDATE email_tasks
                SET status = ?,
                    source = ?,
                    match_score = NULL,
                    match_reason = NULL,
                    fit_points = NULL,
                    risk_points = NULL,
                    match_keywords = NULL,
                    generated_subject = NULL,
                    generated_content_text = NULL,
                    generated_content_html = NULL,
                    approved_subject = NULL,
                    approved_body_text = NULL,
                    approved_body_html = NULL,
                    sent_at = CURRENT_TIMESTAMP
                WHERE id = ?
                """,
                ("sent", "manual", parent_task_id),
            )
            connection.commit()
        finally:
            connection.close()

        response = self.client.post(f"/api/email-tasks/{parent_task_id}/start-follow-up")

        self.assertEqual(response.status_code, 200, msg=response.text)
        current_task = response.json()["current_task"]
        self.assertEqual(current_task["status"], "matched")
        self.assertEqual(current_task["parent_task_id"], parent_task_id)
        self.assertIsNone(current_task["match_score"])
        self.assertIsNone(current_task["match_reason"])

    def test_start_follow_up_creates_manual_child_task_from_reply_detected_task(self) -> None:
        identity_id = self._create_identity(with_imap=False)
        llm_id = self._create_llm()
        primary_material_id = self._upload_material(
            identity_id,
            filename="resume.txt",
            content=b"My background covers agent systems and information extraction.",
            material_type="resume",
        )
        set_primary_response = self.client.post(f"/api/materials/{primary_material_id}/set-primary")
        self.assertEqual(set_primary_response.status_code, 200, msg=set_primary_response.text)

        professor_response = self.client.post(
            "/api/professors",
            json={
                "name": "回复后跟进导师",
                "email": "follow-up-replied@example.edu",
                "title": "Professor",
                "university": "Example University",
                "school": "School of Computing",
                "department": "Computer Science",
                "research_direction": "Information extraction",
                "recent_papers": [],
                "profile_url": None,
                "source_url": None,
            },
        )
        self.assertEqual(professor_response.status_code, 201, msg=professor_response.text)
        professor_id = professor_response.json()["id"]

        ensure_response = self.client.post(
            f"/api/workspaces/{professor_id}/ensure-task",
            params={"identity_id": identity_id, "llm_profile_id": llm_id},
        )
        self.assertEqual(ensure_response.status_code, 200, msg=ensure_response.text)
        parent_task_id = ensure_response.json()["current_task"]["id"]

        connection = sqlite3.connect(self.db_path)
        try:
            connection.execute(
                """
                UPDATE email_tasks
                SET status = ?,
                    source = ?,
                    primary_material_id = ?,
                    match_score = ?,
                    match_reason = ?,
                    fit_points = ?,
                    risk_points = ?,
                    match_keywords = ?,
                    generated_subject = ?,
                    generated_content_text = ?,
                    generated_content_html = ?,
                    approved_subject = ?,
                    approved_body_text = ?,
                    approved_body_html = ?,
                    is_replied = 1,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
                """,
                (
                    "reply_detected",
                    "manual",
                    primary_material_id,
                    86,
                    "对方已回复，适合继续跟进",
                    json.dumps(["已建立对话"]),
                    json.dumps(["需明确下一步诉求"]),
                    json.dumps(["reply"]),
                    "旧跟进草稿主题",
                    "旧跟进草稿正文",
                    "<p>旧跟进草稿正文</p>",
                    "旧审批主题",
                    "旧审批正文",
                    "<p>旧审批正文</p>",
                    parent_task_id,
                ),
            )
            connection.commit()
        finally:
            connection.close()

        workspace_before = self.client.get(
            f"/api/workspaces/{professor_id}",
            params={"identity_id": identity_id, "llm_profile_id": llm_id},
        )
        self.assertEqual(workspace_before.status_code, 200, msg=workspace_before.text)
        self.assertTrue(workspace_before.json()["current_task"]["can_write_follow_up"])

        response = self.client.post(f"/api/email-tasks/{parent_task_id}/start-follow-up")

        self.assertEqual(response.status_code, 200, msg=response.text)
        current_task = response.json()["current_task"]
        self.assertEqual(current_task["parent_task_id"], parent_task_id)
        self.assertEqual(current_task["source"], "manual")
        self.assertEqual(current_task["status"], "matched")
        self.assertIsNone(current_task["generated_subject"])
        self.assertIsNone(current_task["generated_content_text"])
        self.assertIsNone(current_task["approved_subject"])
        self.assertIsNone(current_task["approved_body_text"])
        self.assertEqual(current_task["match_score"], 86)
        self.assertEqual(current_task["match_reason"], "对方已回复，适合继续跟进")

    def test_start_follow_up_rejects_duplicate_manual_child_creation(self) -> None:
        identity_id = self._create_identity(with_imap=False)
        llm_id = self._create_llm()

        professor_response = self.client.post(
            "/api/professors",
            json={
                "name": "跟进重复派生导师",
                "email": "follow-up-duplicate@example.edu",
                "title": "Professor",
                "university": "Example University",
                "school": "School of Computing",
                "department": "Computer Science",
                "research_direction": "Information extraction",
                "recent_papers": [],
                "profile_url": None,
                "source_url": None,
            },
        )
        self.assertEqual(professor_response.status_code, 201, msg=professor_response.text)
        professor_id = professor_response.json()["id"]

        ensure_response = self.client.post(
            f"/api/workspaces/{professor_id}/ensure-task",
            params={"identity_id": identity_id, "llm_profile_id": llm_id},
        )
        self.assertEqual(ensure_response.status_code, 200, msg=ensure_response.text)
        parent_task_id = ensure_response.json()["current_task"]["id"]

        connection = sqlite3.connect(self.db_path)
        try:
            connection.execute(
                """
                UPDATE email_tasks
                SET status = ?, source = ?, sent_at = CURRENT_TIMESTAMP
                WHERE id = ?
                """,
                ("sent", "manual", parent_task_id),
            )
            connection.commit()
        finally:
            connection.close()

        first_response = self.client.post(f"/api/email-tasks/{parent_task_id}/start-follow-up")
        second_response = self.client.post(f"/api/email-tasks/{parent_task_id}/start-follow-up")

        self.assertEqual(first_response.status_code, 200, msg=first_response.text)
        self.assertEqual(second_response.status_code, 400, msg=second_response.text)

    def test_start_follow_up_returns_404_for_missing_task(self) -> None:
        response = self.client.post("/api/email-tasks/9999/start-follow-up")

        self.assertEqual(response.status_code, 404, msg=response.text)
        self.assertEqual(response.json()["detail"], "EmailTask 9999 不存在")

    def test_start_follow_up_rejects_task_without_sent_or_reply_detected_guard(self) -> None:
        identity_id = self._create_identity(with_imap=False)
        llm_id = self._create_llm()

        professor_response = self.client.post(
            "/api/professors",
            json={
                "name": "跟进非法状态导师",
                "email": "follow-up-guard@example.edu",
                "title": "Professor",
                "university": "Example University",
                "school": "School of Computing",
                "department": "Computer Science",
                "research_direction": "Information extraction",
                "recent_papers": [],
                "profile_url": None,
                "source_url": None,
            },
        )
        self.assertEqual(professor_response.status_code, 201, msg=professor_response.text)
        professor_id = professor_response.json()["id"]

        ensure_response = self.client.post(
            f"/api/workspaces/{professor_id}/ensure-task",
            params={"identity_id": identity_id, "llm_profile_id": llm_id},
        )
        self.assertEqual(ensure_response.status_code, 200, msg=ensure_response.text)
        task_id = ensure_response.json()["current_task"]["id"]

        connection = sqlite3.connect(self.db_path)
        try:
            connection.execute(
                """
                UPDATE email_tasks
                SET status = ?
                WHERE id = ?
                """,
                ("matched", task_id),
            )
            connection.commit()
        finally:
            connection.close()

        response = self.client.post(f"/api/email-tasks/{task_id}/start-follow-up")

        self.assertEqual(response.status_code, 400, msg=response.text)

    def test_batch_task_without_default_material_can_still_send_manually(self) -> None:
        identity_id = self._create_identity(with_imap=False)
        llm_id = self._create_llm()

        self.client.post("/api/professors/import-sample")
        professor_id = self.client.get("/api/professors").json()[0]["id"]

        response = self.client.post(
            "/api/batch-tasks",
            json={
                "identity_id": identity_id,
                "llm_profile_id": llm_id,
                "name": "无默认材料也可创建",
                "professor_ids": [professor_id],
                "schedule_type": "immediate",
                "window_start_time": None,
                "window_end_time": None,
                "emails_per_window": None,
                "primary_material_id": None,
                "email_subject": "申请与{{name}}老师交流",
                "email_body": "老师您好，我是{{sender_name}}，希望与您进一步交流。",
                "selected_material_ids": None,
            },
        )

        self.assertEqual(response.status_code, 201)

        workspace = self.client.get(
            f"/api/workspaces/{professor_id}",
            params={"identity_id": identity_id, "llm_profile_id": llm_id},
        )
        task_id = workspace.json()["current_task"]["id"]
        self.assertIsNone(workspace.json()["current_task"]["primary_material_id"])

        regenerate_response = self.client.post(f"/api/email-tasks/{task_id}/generate-draft")
        self.assertEqual(regenerate_response.status_code, 400)
        self.assertEqual(regenerate_response.json()["detail"], "请先选择用于匹配的默认材料")

        with patch(
            "app.services.task_runtime.mail_runtime.send_email",
            AsyncMock(
                return_value=self._build_send_result(
                    message_id="<manual-no-primary@example.com>",
                    provider_payload={"smtp_host": "smtp.example.com"},
                ),
            ),
        ):
            send_response = self.client.post(
                f"/api/email-tasks/{task_id}/approve-and-send",
                json={
                    "subject": "手动邮件",
                    "body_text": "老师您好，这是一封手动编写的邮件。",
                    "body_html": None,
                    "selected_material_ids": [],
                },
            )
        self.assertEqual(send_response.status_code, 200)
        self.assertEqual(send_response.json()["current_task"]["status"], "sent")

    def test_template_polish_mode_requires_complete_template_when_creating_batch_task(self) -> None:
        identity_id = self._create_identity(with_imap=False)
        llm_id = self._create_llm()
        material_id = self._upload_material(
            identity_id,
            filename="resume.txt",
            content=b"My background covers information extraction.",
            material_type="resume",
        )
        connection = sqlite3.connect(self.db_path)
        try:
            connection.execute(
                """
                UPDATE identity_profiles
                SET outreach_template_subject = NULL,
                    outreach_template_body_text = NULL,
                    outreach_template_body_html = NULL
                WHERE id = ?
                """,
                (identity_id,),
            )
            connection.commit()
        finally:
            connection.close()
        self.client.post("/api/professors/import-sample")
        professor_id = self.client.get("/api/professors").json()[0]["id"]

        response = self.client.post(
            "/api/batch-tasks",
            json={
                "identity_id": identity_id,
                "llm_profile_id": llm_id,
                "name": "模板润色缺模板",
                "professor_ids": [professor_id],
                "schedule_type": "immediate",
                "window_start_time": None,
                "window_end_time": None,
                "emails_per_window": None,
                "primary_material_id": material_id,
                "email_subject": None,
                "email_body": None,
                "selected_material_ids": None,
                "outreach_generation_mode": "llm",
                "outreach_template_subject": None,
                "outreach_template_body_text": None,
                "outreach_template_body_html": None,
            },
        )

        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.json()["detail"], "请先填写默认套磁信主题和纯文本正文")

    def test_llm_mode_requires_complete_template_before_generating_draft(self) -> None:
        identity_id = self._create_identity(with_imap=False)
        llm_id = self._create_llm()
        self._upload_material(
            identity_id,
            filename="resume.txt",
            content=b"My research focuses on information extraction and agents.",
            material_type="resume",
        )
        self.client.post("/api/professors/import-sample")
        professor_id = self.client.get("/api/professors").json()[0]["id"]

        ensure_response = self.client.post(
            f"/api/workspaces/{professor_id}/ensure-task",
            params={"identity_id": identity_id, "llm_profile_id": llm_id},
        )
        self.assertEqual(ensure_response.status_code, 200, msg=ensure_response.text)
        task_id = ensure_response.json()["current_task"]["id"]
        connection = sqlite3.connect(self.db_path)
        try:
            connection.execute(
                """
                UPDATE identity_profiles
                SET outreach_template_subject = NULL,
                    outreach_template_body_text = NULL,
                    outreach_template_body_html = NULL
                WHERE id = ?
                """,
                (identity_id,),
            )
            connection.execute(
                """
                UPDATE email_tasks
                SET outreach_generation_mode = ?, outreach_template_subject = NULL,
                    outreach_template_body_text = NULL, outreach_template_body_html = NULL
                WHERE id = ?
                """,
                ("llm", task_id),
            )
            connection.commit()
        finally:
            connection.close()

        with patch(
            "app.services.task_runtime.llm_runtime.generate_draft_content",
            AsyncMock(
                return_value=self._build_draft_generation_result(
                    subject="测试草稿",
                    body_text="测试正文",
                    body_html="<p>测试正文</p>",
                ),
            ),
        ) as mocked_generate:
            generate_response = self.client.post(f"/api/email-tasks/{task_id}/generate-draft")

        self.assertEqual(generate_response.status_code, 400)
        self.assertEqual(generate_response.json()["detail"], "请先填写默认套磁信主题和纯文本正文")
        mocked_generate.assert_not_awaited()

    def test_schedule_and_reply_detection_use_sent_message_id(self) -> None:
        identity_id = self._create_identity(with_imap=True)
        llm_id = self._create_llm()
        self.client.post("/api/professors/import-sample")
        professor_id = self.client.get("/api/professors").json()[0]["id"]
        self.client.post(
            "/api/batch-tasks",
            json={
                "identity_id": identity_id,
                "llm_profile_id": llm_id,
                "name": "回信跟踪任务",
                "professor_ids": [professor_id],
                "schedule_type": "immediate",
                "window_start_time": None,
                "window_end_time": None,
                "emails_per_window": None,
                "primary_material_id": None,
                "email_subject": "申请与{{name}}老师交流",
                "email_body": "老师您好，我是{{sender_name}}，后续会手动整理并发送这封邮件。",
                "selected_material_ids": None,
            },
        )

        workspace = self.client.get(
            f"/api/workspaces/{professor_id}",
            params={"identity_id": identity_id, "llm_profile_id": llm_id},
        ).json()
        task_id = workspace["current_task"]["id"]

        schedule_time = datetime.now(UTC) + timedelta(hours=1)
        with patch(
            "app.services.task_runtime.mail_runtime.send_email",
            AsyncMock(
                return_value=self._build_send_result(
                    message_id="<msg-1@example.com>",
                    provider_payload={"smtp_host": "smtp.example.com"},
                ),
            ),
        ):
            schedule_response = self.client.post(
                f"/api/email-tasks/{task_id}/approve-and-schedule",
                json={
                    "subject": "套磁申请",
                    "body_text": "老师您好，我计划在稍后发送这封邮件。",
                    "body_html": None,
                    "selected_material_ids": [],
                    "scheduled_at": schedule_time.isoformat(),
                },
            )
            self.assertEqual(schedule_response.status_code, 200)

            self._run_async(self._force_task_due(task_id))
            self._run_async(self._dispatch_due_tasks())

        sent_workspace = self.client.get(
            f"/api/workspaces/{professor_id}",
            params={"identity_id": identity_id, "llm_profile_id": llm_id},
        ).json()
        self.assertEqual(sent_workspace["current_task"]["status"], "sent")
        self.assertEqual(sent_workspace["current_task"]["last_rfc_message_id"], "<msg-1@example.com>")

        with patch(
            "app.services.task_runtime.mail_runtime.fetch_recent_inbox_messages",
            AsyncMock(
                return_value=[
                    self._build_received_email(
                        from_email="sample.professor@example.edu",
                        subject="Re: 套磁申请",
                        content="谢谢来信，我们可以进一步聊聊。",
                        message_id="<reply-1@example.com>",
                        in_reply_to="<msg-1@example.com>",
                    ),
                ],
            ),
        ):
            self._run_async(self._poll_replies())

        replied_workspace = self.client.get(
            f"/api/workspaces/{professor_id}",
            params={"identity_id": identity_id, "llm_profile_id": llm_id},
        ).json()
        self.assertEqual(replied_workspace["current_task"]["status"], "reply_detected")
        self.assertTrue(replied_workspace["current_task"]["is_replied"])
        self.assertEqual(replied_workspace["messages"][-1]["direction"], "received")

    def test_batch_task_card_hides_delivery_mode_snapshot(self) -> None:
        identity_id = self._create_identity(with_imap=False)
        llm_id = self._create_llm()
        self.client.post("/api/professors/import-sample")
        professor_id = self.client.get("/api/professors").json()[0]["id"]

        create_response = self.client.post(
            "/api/batch-tasks",
            json={
                "identity_id": identity_id,
                "llm_profile_id": llm_id,
                "name": "批量列表不再显示模式",
                "professor_ids": [professor_id],
                "schedule_type": "immediate",
                "window_start_time": None,
                "window_end_time": None,
                "emails_per_window": None,
                "primary_material_id": None,
                "email_subject": "申请与{{name}}老师交流",
                "email_body": "老师您好，我是{{sender_name}}。",
                "selected_material_ids": None,
            },
        )

        self.assertEqual(create_response.status_code, 201, msg=create_response.text)

        task_payload = self.client.get(
            "/api/batch-tasks",
            params={"identity_id": identity_id, "llm_profile_id": llm_id},
        ).json()[0]
        self.assertNotIn("dry_run_count", task_payload)
        self.assertNotIn("live_count", task_payload)

    def test_batch_task_card_counts_draft_generation_statuses(self) -> None:
        identity_id = self._create_identity(with_imap=False)
        llm_id = self._create_llm()
        self.client.post("/api/professors/import-sample")
        professors = self.client.get("/api/professors").json()[:4]

        create_response = self.client.post(
            "/api/batch-tasks",
            json={
                "identity_id": identity_id,
                "llm_profile_id": llm_id,
                "name": "草稿统计任务",
                "professor_ids": [item["id"] for item in professors],
                "schedule_type": "immediate",
                "window_start_time": None,
                "window_end_time": None,
                "emails_per_window": None,
                "primary_material_id": None,
                "email_subject": "申请与{{name}}老师交流",
                "email_body": "老师您好，我是{{sender_name}}。",
                "selected_material_ids": None,
            },
        )
        self.assertEqual(create_response.status_code, 201, msg=create_response.text)
        batch_task_id = create_response.json()["id"]

        connection = sqlite3.connect(self.db_path)
        try:
            task_ids = [
                row[0]
                for row in connection.execute(
                    """
                    SELECT id
                    FROM email_tasks
                    WHERE batch_task_id = ?
                    ORDER BY id
                    """,
                    (batch_task_id,),
                ).fetchall()
            ]
            self.assertEqual(len(task_ids), 4)
            connection.execute("UPDATE email_tasks SET status = 'generating_draft' WHERE id = ?", (task_ids[0],))
            connection.execute("UPDATE email_tasks SET status = 'draft_failed' WHERE id = ?", (task_ids[1],))
            connection.execute("UPDATE email_tasks SET status = 'review_required' WHERE id = ?", (task_ids[2],))
            connection.commit()
        finally:
            connection.close()

        response = self.client.get("/api/batch-tasks")
        self.assertEqual(response.status_code, 200, msg=response.text)
        task_payload = next(item for item in response.json() if item["id"] == batch_task_id)
        self.assertEqual(task_payload["generating_draft_count"], 1)
        self.assertEqual(task_payload["draft_failed_count"], 1)
        self.assertEqual(task_payload["pending_generation_count"], 1)

    def test_template_batch_task_creates_approved_items_without_review(self) -> None:
        identity_id = self._create_identity(with_imap=False)
        llm_id = self._create_llm()
        professor_response = self.client.post(
            "/api/professors",
            json={
                "name": "模板直通导师",
                "email": "template-direct@example.edu",
                "title": "Professor",
                "university": "Example University",
                "school": "School of Computing",
                "department": "Computer Science",
                "research_direction": "Agents",
                "recent_papers": [],
                "profile_url": None,
                "source_url": None,
            },
        )
        self.assertEqual(professor_response.status_code, 201, msg=professor_response.text)
        professor_id = professor_response.json()["id"]

        response = self.client.post(
            "/api/batch-tasks",
            json={
                "identity_id": identity_id,
                "llm_profile_id": llm_id,
                "name": "模板批量任务",
                "professor_ids": [professor_id],
                "schedule_type": "immediate",
                "window_start_time": None,
                "window_end_time": None,
                "emails_per_window": None,
                "primary_material_id": None,
                "email_subject": None,
                "email_body": None,
                "selected_material_ids": None,
                "outreach_generation_mode": "template",
                "outreach_template_subject": "发送给{{name}}",
                "outreach_template_body_text": "{{name}}老师您好，我是{{sender_name}}。",
                "outreach_template_body_html": "<p>{{name}}老师您好，我是{{sender_name}}。</p>",
            },
        )

        self.assertEqual(response.status_code, 201, msg=response.text)
        task_id = response.json()["id"]
        items = self.client.get(f"/api/batch-tasks/{task_id}/items")
        self.assertEqual(items.status_code, 200, msg=items.text)
        self.assertEqual(items.json()[0]["status"], "approved")

    def test_batch_task_items_show_professor_delivery_progress(self) -> None:
        identity_id = self._create_identity(with_imap=False)
        llm_id = self._create_llm()
        self.client.post("/api/professors/import-sample")
        professors = self.client.get("/api/professors").json()[:2]

        create_response = self.client.post(
            "/api/batch-tasks",
            json={
                "identity_id": identity_id,
                "llm_profile_id": llm_id,
                "name": "明细任务",
                "professor_ids": [item["id"] for item in professors],
                "schedule_type": "immediate",
                "window_start_time": None,
                "window_end_time": None,
                "emails_per_window": None,
                "primary_material_id": None,
                "email_subject": "申请与{{name}}老师交流",
                "email_body": "老师您好，我是{{sender_name}}。",
                "selected_material_ids": None,
            },
        )
        self.assertEqual(create_response.status_code, 201, msg=create_response.text)
        batch_task_id = create_response.json()["id"]

        connection = sqlite3.connect(self.db_path)
        try:
            task_ids = [
                row[0]
                for row in connection.execute(
                    """
                    SELECT id
                    FROM email_tasks
                    WHERE batch_task_id = ?
                    ORDER BY id
                    """,
                    (batch_task_id,),
                ).fetchall()
            ]
            self.assertEqual(len(task_ids), 2)
            connection.execute(
                """
                UPDATE email_tasks
                SET status = 'sent', sent_at = CURRENT_TIMESTAMP
                WHERE id = ?
                """,
                (task_ids[0],),
            )
            connection.execute(
                """
                UPDATE email_tasks
                SET status = 'send_failed', last_error = 'smtp timeout'
                WHERE id = ?
                """,
                (task_ids[1],),
            )
            connection.commit()
        finally:
            connection.close()

        response = self.client.get(f"/api/batch-tasks/{batch_task_id}/items")
        self.assertEqual(response.status_code, 200, msg=response.text)
        payload = response.json()
        self.assertEqual(len(payload), 2)
        self.assertEqual(payload[0]["professor_id"], professors[0]["id"])
        self.assertEqual(payload[0]["professor_name"], professors[0]["name"])
        self.assertEqual(payload[0]["status"], "sent")
        self.assertIsNotNone(payload[0]["sent_at"])
        self.assertEqual(payload[1]["status"], "send_failed")
        self.assertEqual(payload[1]["last_error"], "smtp timeout")

    def test_test_compose_page_can_generate_and_send_to_self(self) -> None:
        identity_id = self._create_identity(with_imap=False)
        llm_id = self._create_llm()
        material_id = self._upload_material(
            identity_id,
            filename="resume.txt",
            content=b"My research focuses on information extraction and agents.",
            material_type="resume",
        )
        set_primary_response = self.client.post(f"/api/materials/{material_id}/set-primary")
        self.assertEqual(set_primary_response.status_code, 200, msg=set_primary_response.text)

        with (
            patch(
                "app.services.test_compose_runtime.llm_runtime.generate_draft_content",
                AsyncMock(
                    return_value=self._build_draft_generation_result(
                        subject="测试主题",
                        body_text="测试正文",
                        body_html="<p>测试正文</p>",
                    ),
                ),
            ),
            patch(
                "app.services.test_compose_runtime.mail_runtime.send_email_to_recipient",
                AsyncMock(
                    return_value=self._build_send_result(
                        message_id="<self-test@example.com>",
                        provider_payload={"to": "sender@example.com"},
                    ),
                ),
            ) as mocked_send,
        ):
            thread_response = self.client.get(f"/api/test-compose/{identity_id}/{llm_id}")
            draft_response = self.client.post(f"/api/test-compose/{identity_id}/{llm_id}/generate-draft")
            send_response = self.client.post(
                f"/api/test-compose/{identity_id}/{llm_id}/send",
                json={
                    "subject": "测试主题",
                    "body_text": "测试正文",
                    "body_html": "<p>测试正文</p>",
                    "selected_material_ids": [],
                },
            )

        self.assertEqual(thread_response.status_code, 200, msg=thread_response.text)
        self.assertEqual(draft_response.status_code, 200, msg=draft_response.text)
        self.assertEqual(send_response.status_code, 200, msg=send_response.text)

        thread_payload = thread_response.json()
        draft_payload = draft_response.json()
        send_payload = send_response.json()

        self.assertEqual(thread_payload["draft"]["selected_material_ids"], [])
        self.assertEqual(draft_payload["draft"]["subject"], "测试主题")
        self.assertEqual(draft_payload["draft"]["body_text"], "测试正文")
        self.assertEqual(send_payload["history"][0]["recipient_email"], "sender@example.com")
        self.assertEqual(send_payload["history"][0]["status"], "sent")
        self.assertEqual(send_payload["history"][0]["rfc_message_id"], "<self-test@example.com>")
        mocked_send.assert_awaited_once()

    def test_test_compose_generate_draft_returns_bad_gateway_when_llm_fails(self) -> None:
        identity_id = self._create_identity(with_imap=False)
        llm_id = self._create_llm()
        material_id = self._upload_material(
            identity_id,
            filename="resume.txt",
            content=b"My research focuses on information extraction and agents.",
            material_type="resume",
        )
        set_primary_response = self.client.post(f"/api/materials/{material_id}/set-primary")
        self.assertEqual(set_primary_response.status_code, 200, msg=set_primary_response.text)

        from app.services import llm_runtime

        with patch(
            "app.services.test_compose_runtime.llm_runtime.generate_draft_content",
            AsyncMock(
                side_effect=llm_runtime.LLMRuntimeError(
                    "模型返回的正文无效",
                    endpoint_kind="chat_completions",
                    status_code=500,
                ),
            ),
        ):
            response = self.client.post(f"/api/test-compose/{identity_id}/{llm_id}/generate-draft")

        self.assertEqual(response.status_code, 502, msg=response.text)
        self.assertEqual(response.json()["detail"], "模型返回的正文无效")

    def test_test_compose_status_is_completed_by_identity_across_llm_profiles(self) -> None:
        identity_id = self._create_identity(with_imap=False)
        first_llm_id = self._create_llm()
        second_llm_response = self.client.post(
            "/api/llm-profiles",
            json={
                "name": "备用模型",
                "provider": "openai",
                "api_base_url": "https://api-backup.example.com/v1",
                "api_key": "sk-test-backup",
                "model_name": "gpt-backup",
                "matcher_prompt_template": "matcher",
                "writer_prompt_template": "writer",
                "temperature": 0.2,
                "max_tokens": 2048,
                "is_default": False,
            },
        )
        self.assertEqual(second_llm_response.status_code, 201, msg=second_llm_response.text)

        with patch(
            "app.services.test_compose_runtime.mail_runtime.send_email_to_recipient",
            AsyncMock(
                return_value=self._build_send_result(
                    message_id="<identity-status@example.com>",
                    provider_payload={},
                ),
            ),
        ):
            send_response = self.client.post(
                f"/api/test-compose/{identity_id}/{first_llm_id}/send",
                json={
                    "subject": "测试主题",
                    "body_text": "测试正文",
                    "body_html": "<p>测试正文</p>",
                    "selected_material_ids": [],
                },
            )

        self.assertEqual(send_response.status_code, 200, msg=send_response.text)

        status_response = self.client.get(f"/api/test-compose/{identity_id}/status")

        self.assertEqual(status_response.status_code, 200, msg=status_response.text)
        self.assertTrue(status_response.json()["completed"])

    def test_test_compose_template_generation_preserves_placeholders_in_draft(self) -> None:
        response = self.client.post(
            "/api/identities",
            json=self._build_identity_payload(
                with_imap=False,
                outreach_generation_mode="template",
                outreach_template_subject="测试给{{name}}",
                outreach_template_body_text="{{name}}您好，我是{{sender_name}}。",
                outreach_template_body_html="<p>{{name}}您好，我是{{sender_name}}。</p>",
            ),
        )
        self.assertEqual(response.status_code, 201, msg=response.text)
        identity_id = response.json()["id"]
        llm_id = self._create_llm()

        draft_response = self.client.post(f"/api/test-compose/{identity_id}/{llm_id}/generate-draft")

        self.assertEqual(draft_response.status_code, 200, msg=draft_response.text)
        draft = draft_response.json()["draft"]
        self.assertEqual(draft["subject"], "测试给{{name}}")
        self.assertIn("{{name}}您好", draft["body_text"])
        self.assertIn("{{sender_name}}", draft["body_text"])
        self.assertIn("{{name}}您好", draft["body_html"])
        self.assertIn("{{sender_name}}", draft["body_html"])

    def test_test_compose_send_renders_placeholders_before_sending(self) -> None:
        identity_id = self._create_identity(with_imap=False)
        llm_id = self._create_llm()

        update_payload = self._build_identity_payload(
            with_imap=False,
            outreach_template_subject="测试给{{name}}",
            outreach_template_body_text="{{name}}您好，我是{{sender_name}}。",
            outreach_template_body_html="<p>{{name}}您好，我是{{sender_name}}。</p>",
        )
        update_payload["profile_name"] = "测试配置"
        update_payload["sender_name"] = "王同学"
        self.client.put(f"/api/identities/{identity_id}", json=update_payload)

        with patch(
            "app.services.test_compose_runtime.mail_runtime.send_email_to_recipient",
            AsyncMock(return_value=self._build_send_result(message_id="<test-render@example.com>", provider_payload={})),
        ) as mocked_send:
            response = self.client.post(
                f"/api/test-compose/{identity_id}/{llm_id}/send",
                json={
                    "subject": "发送给{{name}}",
                    "body_text": "{{name}}您好，我是{{sender_name}}，研究方向：{{research_direction}}。",
                    "body_html": "<p>{{name}}您好，我是{{sender_name}}，研究方向：{{research_direction}}。</p>",
                    "selected_material_ids": [],
                },
            )

        self.assertEqual(response.status_code, 200, msg=response.text)
        kwargs = mocked_send.await_args.kwargs
        self.assertEqual(kwargs["recipient_name"], "测试收件人")
        self.assertEqual(kwargs["subject"], "发送给测试收件人")
        self.assertIn("测试收件人您好", kwargs["body_text"])
        self.assertIn("我是王同学", kwargs["body_text"])
        self.assertIn("测试研究方向", kwargs["body_text"])
        self.assertNotIn("{{name}}", kwargs["body_html"])

        history = response.json()["history"][0]
        self.assertEqual(history["subject"], "发送给测试收件人")
        self.assertIn("测试收件人您好", history["content"])
        self.assertNotIn("{{sender_name}}", history["content_html"])

        draft = response.json()["draft"]
        self.assertEqual(draft["subject"], "发送给{{name}}")
        self.assertIn("{{name}}您好", draft["body_text"])
        self.assertIn("{{sender_name}}", draft["body_text"])
        self.assertIn("{{name}}您好", draft["body_html"])
        self.assertIn("{{sender_name}}", draft["body_html"])

    def test_batch_task_outreach_snapshot_is_independent_from_identity_defaults(self) -> None:
        identity_id = self._create_identity(with_imap=False)
        llm_id = self._create_llm()

        self.client.put(
            f"/api/identities/{identity_id}",
            json=self._build_identity_payload(
                with_imap=False,
                outreach_generation_mode="llm",
                outreach_template_subject="默认主题 {{name}}",
                outreach_template_body_text="默认正文 {{name}}",
                outreach_template_body_html="<p>默认正文 {{name}}</p>",
            ),
        )

        professor_response = self.client.post(
            "/api/professors",
            json={
                "name": "批量快照导师",
                "email": "batch-snapshot@example.edu",
                "title": "Professor",
                "university": "Example University",
                "school": "School of Computing",
                "department": "Computer Science",
                "research_direction": "Agents",
                "recent_papers": [],
                "profile_url": None,
                "source_url": None,
            },
        )
        professor_id = professor_response.json()["id"]

        response = self.client.post(
            "/api/batch-tasks",
            json={
                "identity_id": identity_id,
                "llm_profile_id": llm_id,
                "name": "批量模板快照",
                "professor_ids": [professor_id],
                "schedule_type": "immediate",
                "window_start_time": None,
                "window_end_time": None,
                "emails_per_window": None,
                "primary_material_id": None,
                "email_subject": None,
                "email_body": None,
                "selected_material_ids": None,
                "outreach_generation_mode": "template",
                "outreach_template_subject": "批量主题 {{name}}",
                "outreach_template_body_text": "批量正文 {{name}}",
                "outreach_template_body_html": "<p>批量正文 {{name}}</p>",
            },
        )
        self.assertEqual(response.status_code, 201, msg=response.text)

        self.client.put(
            f"/api/identities/{identity_id}",
            json=self._build_identity_payload(
                with_imap=False,
                outreach_generation_mode="llm",
                outreach_template_subject="后来改掉的主题",
                outreach_template_body_text="后来改掉的正文 {{name}}",
                outreach_template_body_html="<p>后来改掉的正文 {{name}}</p>",
            ),
        )

        workspace = self.client.get(
            f"/api/workspaces/{professor_id}",
            params={"identity_id": identity_id, "llm_profile_id": llm_id},
        )
        self.assertEqual(workspace.status_code, 200, msg=workspace.text)
        payload = workspace.json()
        self.assertEqual(payload["current_task"]["outreach_generation_mode"], "template")
        self.assertEqual(payload["current_task"]["outreach_template_subject"], "批量主题 {{name}}")
        self.assertEqual(payload["current_task"]["outreach_template_body_text"], "批量正文 {{name}}")

    def test_llm_batch_task_prefers_outreach_template_fields_for_snapshot_and_draft(self) -> None:
        identity_id = self._create_identity(with_imap=False)
        llm_id = self._create_llm()
        material_id = self._upload_material(
            identity_id,
            filename="resume.txt",
            content=b"My background covers agent systems and information extraction.",
            material_type="resume",
        )

        update_response = self.client.put(
            f"/api/identities/{identity_id}",
            json=self._build_identity_payload(
                with_imap=False,
                outreach_generation_mode="llm",
                outreach_template_subject="身份默认主题 {{name}}",
                outreach_template_body_text="身份默认正文 {{name}}",
                outreach_template_body_html="<p>身份默认正文 {{name}}</p>",
            ),
        )
        self.assertEqual(update_response.status_code, 200, msg=update_response.text)

        professor_response = self.client.post(
            "/api/professors",
            json={
                "name": "LLM 批量模板导师",
                "email": "llm-batch-template@example.edu",
                "title": "Professor",
                "university": "Example University",
                "school": "School of Computing",
                "department": "Computer Science",
                "research_direction": "Agents",
                "recent_papers": [],
                "profile_url": None,
                "source_url": None,
            },
        )
        self.assertEqual(professor_response.status_code, 201, msg=professor_response.text)
        professor_id = professor_response.json()["id"]

        batch_subject = "批量润色主题 {{name}}"
        batch_body_text = "批量润色正文 {{name}}"
        batch_body_html = "<p>批量润色正文 {{name}}</p>"

        create_response = self.client.post(
            "/api/batch-tasks",
            json={
                "identity_id": identity_id,
                "llm_profile_id": llm_id,
                "name": "LLM 批量模板快照",
                "professor_ids": [professor_id],
                "schedule_type": "immediate",
                "window_start_time": None,
                "window_end_time": None,
                "emails_per_window": None,
                "primary_material_id": material_id,
                "email_subject": None,
                "email_body": None,
                "selected_material_ids": None,
                "outreach_generation_mode": "llm",
                "outreach_template_subject": batch_subject,
                "outreach_template_body_text": batch_body_text,
                "outreach_template_body_html": batch_body_html,
            },
        )
        self.assertEqual(create_response.status_code, 201, msg=create_response.text)
        self.assertEqual(create_response.json()["email_subject"], batch_subject)

        workspace_before_generate = self.client.get(
            f"/api/workspaces/{professor_id}",
            params={"identity_id": identity_id, "llm_profile_id": llm_id},
        )
        self.assertEqual(workspace_before_generate.status_code, 200, msg=workspace_before_generate.text)
        task_before_generate = workspace_before_generate.json()["current_task"]
        self.assertEqual(task_before_generate["outreach_template_subject"], batch_subject)
        self.assertEqual(task_before_generate["outreach_template_body_text"], batch_body_text)
        self.assertEqual(task_before_generate["outreach_template_body_html"], batch_body_html)

        async def _fake_generate_draft_content(**kwargs):
            self.assertEqual(kwargs["custom_subject"], batch_subject)
            self.assertEqual(kwargs["custom_body"], batch_body_text)
            self.assertEqual(kwargs["max_tokens"], 6000)
            return self._build_draft_generation_result(
                subject=f"润色后: {kwargs['custom_subject']}",
                body_text=f"润色后正文: {kwargs['custom_body']}",
                body_html=f"<p>润色后正文: {kwargs['custom_body']}</p>",
            )

        with patch(
            "app.services.task_runtime.llm_runtime.generate_draft_content",
            AsyncMock(side_effect=_fake_generate_draft_content),
        ) as mocked_generate:
            generate_response = self.client.post(
                f"/api/email-tasks/{task_before_generate['id']}/generate-draft",
            )

        self.assertEqual(generate_response.status_code, 200, msg=generate_response.text)
        generated_task = generate_response.json()["current_task"]
        self.assertEqual(generated_task["generated_subject"], f"润色后: {batch_subject}")
        self.assertEqual(generated_task["generated_content_text"], f"润色后正文: {batch_body_text}")
        mocked_generate.assert_awaited_once()

    def test_identity_missing_returns_utf8_detail_message(self) -> None:
        response = self.client.post("/api/identities/999999/smtp-test")

        self.assertEqual(response.status_code, 404)
        self.assertEqual(response.json()["detail"], "未找到身份配置")

    def test_workspace_mode_switch_uses_task_snapshot_over_identity_defaults(self) -> None:
        identity_id = self._create_identity(with_imap=False)
        llm_id = self._create_llm()

        update_response = self.client.put(
            f"/api/identities/{identity_id}",
            json=self._build_identity_payload(
                with_imap=False,
                outreach_generation_mode="llm",
                outreach_template_subject="申请与{{name}}老师交流",
                outreach_template_body_text="{{name}}老师您好，我是{{sender_name}}。",
                outreach_template_body_html="<p>{{name}}老师您好，我是{{sender_name}}。</p>",
            ),
        )
        self.assertEqual(update_response.status_code, 200, msg=update_response.text)

        professor_response = self.client.post(
            "/api/professors",
            json={
                "name": "工作区切换导师",
                "email": "workspace-mode@example.edu",
                "title": "Professor",
                "university": "Example University",
                "school": "School of Computing",
                "department": "Computer Science",
                "research_direction": "Agents",
                "recent_papers": [],
                "profile_url": None,
                "source_url": None,
            },
        )
        professor_id = professor_response.json()["id"]

        ensure_response = self.client.post(
            f"/api/workspaces/{professor_id}/ensure-task",
            params={"identity_id": identity_id, "llm_profile_id": llm_id},
        )
        self.assertEqual(ensure_response.status_code, 200, msg=ensure_response.text)
        task_id = ensure_response.json()["current_task"]["id"]

        clear_identity_response = self.client.put(
            f"/api/identities/{identity_id}",
            json=self._build_identity_payload(
                with_imap=False,
                outreach_generation_mode="template",
                outreach_template_subject="后来切换成新的默认主题",
                outreach_template_body_text="后来切换成新的默认正文 {{name}}",
                outreach_template_body_html="<p>后来切换成新的默认正文 {{name}}</p>",
            ),
        )
        self.assertEqual(clear_identity_response.status_code, 200, msg=clear_identity_response.text)

        switch_response = self.client.post(
            f"/api/email-tasks/{task_id}/outreach-config",
            json={"outreach_generation_mode": "template"},
        )
        self.assertEqual(switch_response.status_code, 200, msg=switch_response.text)
        self.assertEqual(switch_response.json()["current_task"]["outreach_generation_mode"], "template")

        generate_response = self.client.post(f"/api/email-tasks/{task_id}/generate-draft")
        self.assertEqual(generate_response.status_code, 200, msg=generate_response.text)
        generated = generate_response.json()
        self.assertEqual(generated["current_task"]["status"], "review_required")
        self.assertEqual(generated["current_task"]["generated_subject"], "申请与工作区切换导师老师交流")
        self.assertIn("工作区切换导师老师您好", generated["current_task"]["generated_content_text"])

    def _run_alembic_upgrade(self) -> None:
        env = os.environ.copy()
        result = subprocess.run(
            [sys.executable, "-m", "alembic", "upgrade", "head"],
            cwd=BACKEND_DIR,
            env=env,
            capture_output=True,
            text=True,
            check=False,
        )
        if result.returncode != 0:
            self.fail(
                "Alembic migration failed.\n"
                f"stdout:\n{result.stdout}\n"
                f"stderr:\n{result.stderr}",
            )

    def _create_identity(self, *, with_imap: bool) -> int:
        response = self.client.post(
            "/api/identities",
            json=self._build_identity_payload(
                with_imap=with_imap,
                outreach_template_subject="申请与{{name}}老师交流",
                outreach_template_body_text="老师您好，我是{{sender_name}}，关注到您在{{research_direction}}方向的工作。",
            ),
        )
        self.assertEqual(response.status_code, 201)
        return response.json()["id"]

    @staticmethod
    def _build_identity_payload(
        *,
        with_imap: bool,
        outreach_generation_mode: str = "llm",
        outreach_template_subject: str | None = None,
        outreach_template_body_text: str | None = None,
        outreach_template_body_html: str | None = None,
    ) -> dict[str, object]:
        return {
            "name": "测试身份",
            "email_address": "sender@example.com",
            "smtp_host": "smtp.example.com",
            "smtp_port": 465,
            "smtp_username": "different-login@example.com",
            "smtp_password": "secret",
            "imap_host": "imap.example.com" if with_imap else None,
            "imap_port": 993 if with_imap else None,
            "imap_username": "sender@example.com" if with_imap else None,
            "imap_password": "secret" if with_imap else None,
            "default_language": "zh-CN",
            "outreach_generation_mode": outreach_generation_mode,
            "outreach_template_subject": outreach_template_subject,
            "outreach_template_body_text": outreach_template_body_text,
            "outreach_template_body_html": outreach_template_body_html,
            "match_threshold": None,
            "daily_send_limit": None,
            "send_interval_min": None,
            "send_interval_max": None,
            "same_domain_cooldown_minutes": None,
            "is_default": True,
        }

    def _create_llm(self) -> int:
        response = self.client.post(
            "/api/llm-profiles",
            json={
                "name": "默认模型",
                "provider": "openai",
                "api_base_url": "https://api.example.com/v1",
                "api_key": "sk-test-key",
                "model_name": "gpt-4o-mini",
                "matcher_prompt_template": "matcher",
                "writer_prompt_template": "writer",
                "temperature": 0.2,
                "max_tokens": 2048,
                "is_default": True,
            },
        )
        self.assertEqual(response.status_code, 201)
        return response.json()["id"]

    def _create_canceled_batch_stopped_parent_task(self, *, email: str) -> int:
        identity_id = self._create_identity(with_imap=False)
        llm_id = self._create_llm()

        professor_response = self.client.post(
            "/api/professors",
            json={
                "name": "旧动作拦截导师",
                "email": email,
                "title": "Professor",
                "university": "Example University",
                "school": "School of Computing",
                "department": "Computer Science",
                "research_direction": "Agent systems",
                "recent_papers": [],
                "profile_url": None,
                "source_url": None,
            },
        )
        self.assertEqual(professor_response.status_code, 201, msg=professor_response.text)
        professor_id = professor_response.json()["id"]

        create_response = self.client.post(
            "/api/batch-tasks",
            json={
                "identity_id": identity_id,
                "llm_profile_id": llm_id,
                "name": "旧动作拦截批量任务",
                "professor_ids": [professor_id],
                "schedule_type": "immediate",
                "window_start_time": None,
                "window_end_time": None,
                "emails_per_window": None,
                "primary_material_id": None,
                "email_subject": "联系 {{name}}",
                "email_body": "联系正文 {{name}}",
                "selected_material_ids": None,
            },
        )
        self.assertEqual(create_response.status_code, 201, msg=create_response.text)
        batch_task_id = create_response.json()["id"]

        connection = sqlite3.connect(self.db_path)
        try:
            task_id = connection.execute(
                """
                SELECT id
                FROM email_tasks
                WHERE batch_task_id = ?
                """,
                (batch_task_id,),
            ).fetchone()[0]
            connection.execute(
                """
                UPDATE batch_tasks
                SET status = ?
                WHERE id = ?
                """,
                ("stopped", batch_task_id),
            )
            connection.execute(
                """
                UPDATE email_tasks
                SET status = ?,
                    cancellation_reason = ?
                WHERE id = ?
                """,
                ("canceled", "batch_stopped", task_id),
            )
            connection.commit()
        finally:
            connection.close()

        return task_id

    async def _dispatch_due_tasks(self) -> None:
        from app.core.database import get_session_factory
        from app.services.task_runtime import dispatch_due_tasks_once

        await dispatch_due_tasks_once(get_session_factory(), limit=10)

    async def _poll_replies(self) -> None:
        from app.core.database import get_session_factory
        from app.services.task_runtime import poll_for_replies_once

        await poll_for_replies_once(get_session_factory())

    async def _force_task_due(self, task_id: int) -> None:
        from app.core.database import get_session_factory
        from app.models import EmailTask

        async with get_session_factory()() as session:
            task = await session.get(EmailTask, task_id)
            task.scheduled_at = datetime.now(UTC) - timedelta(minutes=1)
            await session.commit()

    @staticmethod
    def _build_match_evaluation_result(
        *,
        match_score: int,
    ):
        from app.services.llm_runtime import (
            GeneratedMatchEvaluation,
            MatchEvaluationResult,
        )

        return GeneratedMatchEvaluation(
            result=MatchEvaluationResult(
                match_score=match_score,
                match_reason="研究方向和材料内容高度匹配。",
                fit_points=["研究方向一致", "材料信息完整"],
                risk_points=["尚未展开具体合作切口"],
                keywords=["大模型", "信息提取"],
            ),
            usage=None,
        )

    @staticmethod
    def _build_draft_generation_result(
        *,
        subject: str,
        body_text: str,
        body_html: str,
        prompt_tokens: int | None = None,
        completion_tokens: int | None = None,
        cached_tokens: int | None = None,
    ):
        from app.services.llm_runtime import (
            ChatCompletionUsage,
            DraftGenerationResult,
            GeneratedDraftContent,
        )

        return GeneratedDraftContent(
            result=DraftGenerationResult(
                subject=subject,
                body_text=body_text,
                body_html=body_html,
                suggested_material_ids=[],
            ),
            usage=(
                ChatCompletionUsage(
                    prompt_tokens=prompt_tokens,
                    completion_tokens=completion_tokens,
                    total_tokens=(
                        (prompt_tokens or 0) + (completion_tokens or 0)
                        if prompt_tokens is not None and completion_tokens is not None
                        else None
                    ),
                    cached_tokens=cached_tokens,
                )
                if prompt_tokens is not None or completion_tokens is not None
                else None
            ),
        )

    @staticmethod
    def _build_match_and_draft_result(
        *,
        match_score: int,
        subject: str,
        body_text: str,
        body_html: str,
        prompt_tokens: int | None = None,
        completion_tokens: int | None = None,
    ):
        from app.services.llm_runtime import (
            ChatCompletionUsage,
            GeneratedMatchAndDraft,
            MatchAndDraftResult,
        )

        return GeneratedMatchAndDraft(
            result=MatchAndDraftResult(
                match_score=match_score,
                match_reason="研究方向和材料内容高度匹配。",
                fit_points=["研究方向一致", "材料信息完整"],
                risk_points=["尚未展开具体合作切口"],
                keywords=["大模型", "信息抽取"],
                subject=subject,
                body_text=body_text,
                body_html=body_html,
                suggested_material_ids=[],
            ),
            usage=(
                ChatCompletionUsage(
                    prompt_tokens=prompt_tokens,
                    completion_tokens=completion_tokens,
                    total_tokens=(
                        (prompt_tokens or 0) + (completion_tokens or 0)
                        if prompt_tokens is not None and completion_tokens is not None
                        else None
                    ),
                )
                if prompt_tokens is not None or completion_tokens is not None
                else None
            ),
        )

    def _upload_material(
        self,
        identity_id: int,
        *,
        filename: str,
        content: bytes,
        material_type: str,
    ) -> int:
        response = self.client.post(
            f"/api/identities/{identity_id}/materials",
            files={"file": (filename, io.BytesIO(content), "application/octet-stream")},
            data={"material_type": material_type},
        )
        self.assertEqual(response.status_code, 201)
        return response.json()["id"]

    @staticmethod
    def _build_probe_result(*, ok: bool, message: str, resolved_base_url: str, response_preview: str):
        from app.services.llm_runtime import LLMProbeResult

        return LLMProbeResult(
            ok=ok,
            message=message,
            resolved_base_url=resolved_base_url,
            response_preview=response_preview,
        )

    @staticmethod
    def _build_model_catalog_result(
        *,
        ok: bool,
        message: str,
        resolved_base_url: str,
        models: list[str],
        selected_model_available: bool,
    ):
        from app.services.llm_runtime import LLMModelCatalogResult

        return LLMModelCatalogResult(
            ok=ok,
            message=message,
            resolved_base_url=resolved_base_url,
            models=models,
            selected_model_available=selected_model_available,
        )

    def _latest_email_log_provider_payload(self) -> dict[str, object]:
        connection = sqlite3.connect(self.db_path)
        try:
            raw_payload = connection.execute(
                """
                SELECT provider_payload
                FROM email_logs
                WHERE direction = 'draft'
                ORDER BY id DESC
                LIMIT 1
                """
            ).fetchone()[0]
        finally:
            connection.close()
        if isinstance(raw_payload, str):
            parsed = json.loads(raw_payload)
            if isinstance(parsed, dict):
                return parsed
        self.fail("未找到草稿 provider_payload")

    @staticmethod
    def _build_send_result(*, message_id: str, provider_payload: dict[str, str]):
        from app.services.mail_runtime import SendMailResult

        return SendMailResult(message_id=message_id, provider_payload=provider_payload)

    @staticmethod
    def _build_received_email(
        *,
        from_email: str,
        subject: str,
        content: str,
        message_id: str,
        in_reply_to: str,
    ):
        from app.services.mail_runtime import ReceivedEmail

        return ReceivedEmail(
            from_email=from_email,
            subject=subject,
            content=content,
            content_html=None,
            message_id=message_id,
            in_reply_to=in_reply_to,
            references=in_reply_to,
            sent_at=datetime.now(UTC),
            headers={
                "from": from_email,
                "subject": subject,
                "message_id": message_id,
                "in_reply_to": in_reply_to,
                "references": in_reply_to,
                "to": "sender@example.com",
            },
        )

    @staticmethod
    def _run_async(coro):
        return asyncio.run(coro)


if __name__ == "__main__":
    unittest.main()
