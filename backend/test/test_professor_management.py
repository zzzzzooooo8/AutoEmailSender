from __future__ import annotations

import csv
import io
import unittest

from openpyxl import Workbook, load_workbook
from pydantic import ValidationError

from app.models import Professor
from app.schemas.professor import ProfessorUpsertPayload
from app.services.professor_management import (
    PROFESSOR_TEMPLATE_COLUMNS,
    build_professor_export,
    build_professor_template,
    is_valid_professor_email,
    normalize_professor_email,
    normalize_professor_payload,
    parse_professor_import_file,
)


class ProfessorManagementServiceTests(unittest.TestCase):
    def test_email_validation_accepts_common_addresses_and_rejects_invalid_values(self) -> None:
        valid_values = [
            "zhang@example.edu",
            "li.wei+lab@cs.example.edu",
            "  mixed.case@Example.EDU  ",
        ]
        invalid_values = [
            "",
            "plain-address",
            "missing-domain@",
            "@missing-local.example.edu",
            "space in@example.edu",
        ]

        for value in valid_values:
            with self.subTest(value=value):
                self.assertTrue(is_valid_professor_email(value))

        for value in invalid_values:
            with self.subTest(value=value):
                self.assertFalse(is_valid_professor_email(value))

    def test_normalize_professor_email_collapses_obfuscated_domain_dots(self) -> None:
        self.assertEqual(normalize_professor_email("wjchen@sei.ecnu...cn"), "wjchen@sei.ecnu.cn")
        self.assertEqual(normalize_professor_email(" WJCHEN@SEI.ECNU...CN "), "wjchen@sei.ecnu.cn")

    def test_normalize_professor_email_handles_simple_obfuscation_characters(self) -> None:
        cases = {
            "wjchen&#64;sei.ecnu.edu.cn": "wjchen@sei.ecnu.edu.cn",
            "wjchen＠sei．ecnu．edu．cn": "wjchen@sei.ecnu.edu.cn",
            "wjchen\u200b@sei.ecnu.edu.cn": "wjchen@sei.ecnu.edu.cn",
            "wjchen @ sei . ecnu . edu . cn": "wjchen@sei.ecnu.edu.cn",
            "template@example.edu": "template@example.edu",
            "wjchen AT sei DOT ecnu DOT edu DOT cn": "wjchen@sei.ecnu.edu.cn",
            "wjchen[at]sei[dot]ecnu[dot]edu[dot]cn": "wjchen@sei.ecnu.edu.cn",
        }

        for value, expected in cases.items():
            with self.subTest(value=value):
                self.assertEqual(normalize_professor_email(value), expected)

    def test_email_validation_rejects_un_normalized_empty_domain_labels(self) -> None:
        self.assertFalse(is_valid_professor_email("wjchen@sei.ecnu...cn"))

    def test_normalize_professor_payload_trims_name_and_lowercases_email(self) -> None:
        payload = ProfessorUpsertPayload(
            name="  张明远  ",
            email="  ZHANG@EXAMPLE.EDU  ",
            title=" 教授 ",
            university=" 示例大学 ",
            school=" 人工智能学院 ",
            department=" 计算机科学系 ",
            research_direction=" 大语言模型 ",
            recent_papers=" Paper A | Paper B ",
            profile_url=" https://example.edu/zhang ",
            source_url=" https://example.edu/faculty ",
        )

        self.assertEqual(
            normalize_professor_payload(payload),
            {
                "name": "张明远",
                "email": "zhang@example.edu",
                "title": "教授",
                "university": "示例大学",
                "school": "人工智能学院",
                "department": "计算机科学系",
                "research_direction": "大语言模型",
                "recent_papers": ["Paper A", "Paper B"],
                "profile_url": "https://example.edu/zhang",
                "source_url": "https://example.edu/faculty",
            },
        )

    def test_professor_payload_rejects_blank_name_or_email_before_service_validation(self) -> None:
        for field_name, payload in [
            ("name", {"name": " ", "email": "zhang@example.edu"}),
            ("email", {"name": "张明远", "email": " "}),
        ]:
            with self.subTest(field_name=field_name):
                with self.assertRaises(ValidationError):
                    ProfessorUpsertPayload(**payload)

    def test_parse_csv_import_skips_help_and_example_rows_counts_failures_and_deduplicates_by_email(self) -> None:
        csv_content = "\ufeff# 导师导入模板\n".encode("utf-8")
        csv_content += (
            ",".join(PROFESSOR_TEMPLATE_COLUMNS)
            + "\n"
            + "示例：张明远,example@example.edu,教授,示例大学,人工智能学院,计算机科学系,大语言模型,Paper A|Paper B,https://example.edu/zhang,https://example.edu/faculty\n"
            + "张明远,ZHANG@EXAMPLE.EDU,教授,示例大学,人工智能学院,计算机科学系,大语言模型,Paper A| Paper B ,,https://example.edu/faculty\n"
            + "缺邮箱,,教授,示例大学,人工智能学院,计算机科学系,大语言模型,,,,\n"
            + "张明远更新,zhang@example.edu,讲席教授,示例大学,人工智能学院,计算机科学系,智能体,Paper C,https://example.edu/new,https://example.edu/faculty\n"
        ).encode("utf-8")

        parsed = parse_professor_import_file("professors.csv", csv_content)

        self.assertEqual(parsed.failed_count, 1)
        self.assertEqual(list(parsed.data), ["zhang@example.edu"])
        self.assertEqual(
            parsed.data["zhang@example.edu"],
            {
                "name": "张明远更新",
                "email": "zhang@example.edu",
                "title": None,
                "university": "示例大学",
                "school": "人工智能学院",
                "department": "计算机科学系",
                "research_direction": "智能体",
                "recent_papers": ["Paper C"],
                "profile_url": "https://example.edu/new",
                "source_url": "https://example.edu/faculty",
            },
        )

    def test_parse_csv_import_caps_recent_papers_to_first_8(self) -> None:
        csv_content = (
            ",".join(PROFESSOR_TEMPLATE_COLUMNS)
            + "\n"
            + (
                "张三,zhang@example.edu,教授,示例大学,人工智能学院,计算机科学系,大语言模型,"
                "Paper1|Paper2|Paper3|Paper4|Paper5|Paper6|Paper7|Paper8|Paper9|Paper10,,\n"
            )
        ).encode("utf-8-sig")

        parsed = parse_professor_import_file("professors.csv", csv_content)

        self.assertEqual(
            parsed.data["zhang@example.edu"]["recent_papers"],
            ["Paper1", "Paper2", "Paper3", "Paper4", "Paper5", "Paper6", "Paper7", "Paper8"],
        )

    def test_parse_xlsx_import_finds_header_after_help_rows_and_reads_sparse_rows(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["# 帮助说明"])
        sheet.append(PROFESSOR_TEMPLATE_COLUMNS)
        sheet.append(["李伟", "li@example.edu", None, "示例大学"])
        buffer = io.BytesIO()
        workbook.save(buffer)

        parsed = parse_professor_import_file("professors.xlsx", buffer.getvalue())

        self.assertEqual(parsed.failed_count, 0)
        self.assertEqual(parsed.data["li@example.edu"]["name"], "李伟")
        self.assertEqual(parsed.data["li@example.edu"]["university"], "示例大学")
        self.assertEqual(parsed.data["li@example.edu"]["recent_papers"], [])

    def test_parse_import_rejects_unsupported_extension_missing_columns_bad_encoding_and_corrupt_xlsx(self) -> None:
        with self.assertRaisesRegex(ValueError, "仅支持导入 csv 或 xlsx 文件"):
            parse_professor_import_file("professors.txt", b"name,email\n")

        with self.assertRaisesRegex(ValueError, "导入文件缺少必要列"):
            parse_professor_import_file("professors.csv", b"name,email\nzhang,zhang@example.edu\n")

        with self.assertRaisesRegex(ValueError, "CSV 文件请使用 UTF-8 编码"):
            parse_professor_import_file("professors.csv", b"\xff\xfe\x00")

        with self.assertRaisesRegex(ValueError, "XLSX 文件无法读取"):
            parse_professor_import_file("professors.xlsx", b"not an xlsx file")

    def test_build_professor_template_supports_csv_and_xlsx_and_rejects_unknown_format(self) -> None:
        csv_content, csv_media_type, csv_filename = build_professor_template("csv")
        xlsx_content, xlsx_media_type, xlsx_filename = build_professor_template("xlsx")

        self.assertEqual(csv_media_type, "text/csv; charset=utf-8")
        self.assertEqual(csv_filename, "professors_import_template.csv")
        self.assertTrue(csv_content.startswith(b"\xef\xbb\xbf"))
        self.assertIn("name,email,title", csv_content.decode("utf-8-sig"))

        self.assertEqual(
            xlsx_media_type,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertEqual(xlsx_filename, "professors_import_template.xlsx")
        self.assertGreater(len(xlsx_content), 100)

        with self.assertRaisesRegex(ValueError, "仅支持 csv 或 xlsx 模板"):
            build_professor_template("json")

    def test_build_professor_export_csv_can_be_imported_without_changes(self) -> None:
        professor = Professor(
            name="李伟",
            email="li@example.edu",
            title="教授",
            university="示例大学",
            school="人工智能学院",
            department="计算机科学系",
            research_direction="大语言模型",
            recent_papers=["Paper A", "Paper B"],
            profile_url="https://example.edu/li",
            source_url=None,
        )

        content, media_type, filename = build_professor_export([professor], "csv")

        self.assertEqual(media_type, "text/csv; charset=utf-8")
        self.assertEqual(filename, "professors_export.csv")
        self.assertTrue(content.startswith(b"\xef\xbb\xbf"))
        decoded = content.decode("utf-8-sig")
        self.assertIn(",".join(PROFESSOR_TEMPLATE_COLUMNS), decoded)
        self.assertIn("Paper A|Paper B", decoded)
        self.assertNotIn("None", decoded)
        self.assertNotIn("null", decoded)

        parsed = parse_professor_import_file(filename, content)
        self.assertEqual(parsed.failed_count, 0)
        self.assertEqual(parsed.data["li@example.edu"]["name"], "李伟")
        self.assertEqual(parsed.data["li@example.edu"]["recent_papers"], ["Paper A", "Paper B"])

    def test_build_professor_export_xlsx_can_be_imported_without_changes(self) -> None:
        professor = Professor(
            name="王芳",
            email="wang@example.edu",
            title="副教授",
            university="样例大学",
            school="生命科学学院",
            department="生物信息系",
            research_direction="计算生物学",
            recent_papers=["Paper C"],
            profile_url=None,
            source_url="https://example.edu/faculty",
        )

        content, media_type, filename = build_professor_export([professor], "xlsx")

        self.assertEqual(
            media_type,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertEqual(filename, "professors_export.xlsx")
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        rows = list(workbook.active.iter_rows(values_only=True))
        self.assertEqual(list(rows[0]), PROFESSOR_TEMPLATE_COLUMNS)
        self.assertEqual(rows[1][0], "王芳")
        self.assertEqual(rows[1][7], "Paper C")
        self.assertIsNone(rows[1][8])

        parsed = parse_professor_import_file(filename, content)
        self.assertEqual(parsed.failed_count, 0)
        self.assertEqual(parsed.data["wang@example.edu"]["source_url"], "https://example.edu/faculty")

    def test_build_professor_export_empty_file_and_unknown_format(self) -> None:
        csv_content, _, csv_filename = build_professor_export([], "csv")
        parsed_csv = parse_professor_import_file(csv_filename, csv_content)
        self.assertEqual(parsed_csv.failed_count, 0)
        self.assertEqual(parsed_csv.data, {})

        xlsx_content, _, xlsx_filename = build_professor_export([], "xlsx")
        parsed_xlsx = parse_professor_import_file(xlsx_filename, xlsx_content)
        self.assertEqual(parsed_xlsx.failed_count, 0)
        self.assertEqual(parsed_xlsx.data, {})

        with self.assertRaisesRegex(ValueError, "仅支持 csv 或 xlsx 导出"):
            build_professor_export([], "json")

    def test_build_professor_export_escapes_spreadsheet_formulas(self) -> None:
        professor = Professor(
            name="=cmd|' /C calc'!A0",
            email="formula@example.edu",
            title="+教授",
            university="-示例大学",
            school="@人工智能学院",
            department="  =计算机科学系",
            research_direction="大语言模型",
            recent_papers=["=Paper A", "+Paper B", "Normal Paper"],
            profile_url="https://example.edu/formula",
            source_url=None,
        )

        csv_content, _, _ = build_professor_export([professor], "csv")
        csv_rows = list(csv.reader(io.StringIO(csv_content.decode("utf-8-sig"))))
        self.assertEqual(csv_rows[1][0], "'=cmd|' /C calc'!A0")
        self.assertEqual(csv_rows[1][2], "'+教授")
        self.assertEqual(csv_rows[1][3], "'-示例大学")
        self.assertEqual(csv_rows[1][4], "'@人工智能学院")
        self.assertEqual(csv_rows[1][5], "'=计算机科学系")
        self.assertEqual(csv_rows[1][7], "'=Paper A|'+Paper B|Normal Paper")

        xlsx_content, _, _ = build_professor_export([professor], "xlsx")
        workbook = load_workbook(io.BytesIO(xlsx_content), read_only=True, data_only=True)
        rows = list(workbook.active.iter_rows(values_only=True))
        self.assertEqual(rows[1][0], "'=cmd|' /C calc'!A0")
        self.assertEqual(rows[1][2], "'+教授")
        self.assertEqual(rows[1][3], "'-示例大学")
        self.assertEqual(rows[1][4], "'@人工智能学院")
        self.assertEqual(rows[1][5], "'=计算机科学系")
        self.assertEqual(rows[1][7], "'=Paper A|'+Paper B|Normal Paper")


if __name__ == "__main__":
    unittest.main()
